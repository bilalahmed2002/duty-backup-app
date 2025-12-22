"""Playwright automation for NetCHB duty workflow."""

from __future__ import annotations

import asyncio
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Dict, List, Optional, Tuple
from urllib.parse import urljoin

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.async_api import async_playwright, Browser, BrowserContext, Page

from .otp_manager import OTPManager


LOGIN_URL = "https://www.netchb.com/security/"
AMS_SEARCH_URL = "https://www.netchb.com/app/ams/index.jsp"
AMS_SEARCH_POST_URL = "https://www.netchb.com/app/ams/viewMawbs.do"
AMS_MASTER_URL = "https://www.netchb.com/app/ams/mawbMenu.do"
ENTRIES_URL = "https://www.netchb.com/app/entry/index.jsp"
ENTRIES_SEARCH_POST_URL = "https://www.netchb.com/app/entry/processViewEntries.do"
ENTRY_DETAIL_URL = "https://www.netchb.com/app/entry/viewEntry.do"
PRINT7501_URL = "https://www.netchb.com/app/entry/print7501.do"
CUSTOM_REPORT_URL = "https://www.netchb.com/app/entry/customizableReport.jsp"
CUSTOM_REPORT_DOWNLOAD_URL = "https://www.netchb.com/app/entry/downloadCustomizableReport.do"

# 2FA selectors (provided by user)
OTP_INPUT_SELECTOR = "#tfa"
OTP_SUBMIT_SELECTOR = "#tfaForm > div:nth-child(2) > input[type=submit]"
LOGIN_SUCCESS_SELECTOR = "#menuTableBody > tr > td:nth-child(1)"


def _normalize_mawb(mawb: str) -> str:
    digits = "".join(ch for ch in mawb if ch.isdigit())
    if len(digits) != 11:
        raise ValueError(f"MAWB '{mawb}' must contain exactly 11 digits")
    return digits


def _format_excel_date(value) -> str:
    try:
        if isinstance(value, datetime):
            return value.strftime("%m/%d/%y")
        return datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S").strftime("%m/%d/%y")
    except Exception:
        return str(value)


@dataclass
class DutyRunResult:
    mawb: str
    summary: Dict[str, str] = field(default_factory=dict)
    logs: List[str] = field(default_factory=list)
    download_path: Optional[Path] = None

    def to_summary_dict(self) -> Dict[str, str]:
        return self.summary


class NetChbDutyRunner:
    """Wraps Playwright automation for NetCHB duties."""

    def __init__(self, *, headless: bool = True, download_root: Optional[Path] = None) -> None:
        self.headless = headless
        self.download_root = download_root or Path.cwd() / "temp_netchb_downloads"
        self.download_root.mkdir(parents=True, exist_ok=True)
        self.playwright = None
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None
        self.temp_dir = TemporaryDirectory()
        self._logs: List[str] = []
        self._last_entry_rows: Optional[List[Dict]] = None  # Store entry_rows for PDF download reuse

    def log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        entry = f"[{timestamp}] {message}"
        self._logs.append(entry)
        print(entry)  # Also print for immediate visibility

    @property
    def logs(self) -> List[str]:
        return list(self._logs)

    async def __aenter__(self) -> "NetChbDutyRunner":
        await self._setup_browser()
        return self

    async def __aexit__(self, exc_type, exc, tb) -> None:
        await self.cleanup()

    async def _setup_browser(self) -> None:
        import time
        setup_start = time.time()
        
        self.log("STEP 1: Initializing Playwright browser...")
        step_start = time.time()
        self.playwright = await async_playwright().start()
        self.log(f"STEP 1: Playwright initialized ({time.time() - step_start:.2f}s)")
        
        self.log("STEP 2: Launching Chromium browser (headless={})...".format(self.headless))
        step_start = time.time()
        # Use container-safe browser args
        import sys
        from pathlib import Path
        
        # Add app utils to path for imports
        _app_dir = Path(__file__).parent.parent.parent.resolve()
        _utils_dir = _app_dir / "utils"
        if str(_utils_dir) not in sys.path:
            sys.path.insert(0, str(_utils_dir))
        
        from playwright_launcher import get_container_safe_browser_args
        base_args = get_container_safe_browser_args()
        # Add window size for NetCHB
        extra_args = ["--window-size=1920,1080"]
        all_args = base_args + extra_args
        
        self.browser = await self.playwright.chromium.launch(
            headless=self.headless,
            args=all_args,
        )
        self.log(f"STEP 2: Browser launched ({time.time() - step_start:.2f}s)")

        self.log("STEP 3: Creating browser context with download settings...")
        step_start = time.time()
        download_dir = Path(self.temp_dir.name)
        self.context = await self.browser.new_context(
            viewport={"width": 1920, "height": 1080},
            accept_downloads=True,
        )
        self.log(f"STEP 3: Context created ({time.time() - step_start:.2f}s)")

        self.log("STEP 4: Setting up download handling...")
        async def handle_download(download):
            self.log(f"Download started: {download.suggested_filename}")
            path = await download.path()
            self.log(f"Download saved to: {path}")

        self.context.on("download", handle_download)

        self.log("STEP 5: Creating new page...")
        step_start = time.time()
        self.page = await self.context.new_page()
        self.log(f"STEP 5: Page created ({time.time() - step_start:.2f}s)")
        
        total_time = time.time() - setup_start
        self.log(f"✅ Browser setup complete in {total_time:.2f}s. Ready for automation.")

    async def cleanup(self) -> None:
        self.log("CLEANUP: Closing browser and cleaning up...")
        try:
            if self.page:
                await self.page.close()
                self.log("CLEANUP: Page closed")
            if self.context:
                await self.context.close()
                self.log("CLEANUP: Context closed")
            if self.browser:
                await self.browser.close()
                self.log("CLEANUP: Browser closed")
            if self.playwright:
                await self.playwright.stop()
                self.log("CLEANUP: Playwright stopped")
        except Exception as exc:
            self.log(f"CLEANUP ERROR: {exc}")
        finally:
            self.page = None
            self.context = None
            self.browser = None
            self.playwright = None
            self.temp_dir.cleanup()
            self.log("CLEANUP: Complete")

    # ------------------------------------------------------------------
    # Login
    # ------------------------------------------------------------------
    async def login(self, username: str, password: str, otp_uri: Optional[str] = None) -> None:
        if not self.page:
            raise RuntimeError("Page not initialized")

        self.log("=" * 60)
        self.log("LOGIN PROCESS STARTED")
        self.log("=" * 60)

        self.log(f"STEP 1: Navigating to login URL: {LOGIN_URL}")
        await self.page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=30000)
        self.log(f"STEP 1: Page loaded. Current URL: {self.page.url}")

        self.log("STEP 2: Waiting for username field (#lName)...")
        await self.page.wait_for_selector("#lName", timeout=20000)
        self.log("STEP 2: Username field found")

        self.log(f"STEP 3: Filling username field with: {username}")
        await self.page.fill("#lName", username)
        self.log("STEP 3: Username filled")

        self.log("STEP 4: Filling password field (#pass)...")
        await self.page.fill("#pass", password)
        self.log("STEP 4: Password filled")

        self.log("STEP 5: Clicking login submit button...")
        await self.page.click("input[type=submit]")
        self.log("STEP 5: Login button clicked")

        self.log("STEP 6: Waiting for page response (either 2FA or dashboard)...")
        await self.page.wait_for_timeout(2000)  # Give page time to respond

        # Check if 2FA is required
        if otp_uri:
            self.log("=" * 60)
            self.log("2FA AUTHENTICATION REQUIRED")
            self.log("=" * 60)

            self.log("STEP 7: Checking for 2FA input field (#tfa)...")
            try:
                await self.page.wait_for_selector(
                    OTP_INPUT_SELECTOR,
                    timeout=10000,
                    state="visible"
                )
                self.log("STEP 7: 2FA input field found - authentication required")

                self.log("STEP 8: Generating fresh TOTP code from OTP URI...")
                otp_code = OTPManager.get_fresh_otp(otp_uri, min_seconds_remaining=5)
                if not otp_code:
                    raise RuntimeError("Failed to generate TOTP code for 2FA")
                self.log(f"STEP 8: TOTP code generated: {otp_code}")

                self.log("STEP 9: Filling 2FA code into input field...")
                # Playwright's fill() automatically clears the field first
                await self.page.fill(OTP_INPUT_SELECTOR, otp_code)
                self.log("STEP 9: 2FA code filled")

                self.log("STEP 10: Waiting for 2FA submit button...")
                await self.page.wait_for_selector(
                    OTP_SUBMIT_SELECTOR,
                    timeout=5000,
                    state="visible"
                )
                self.log("STEP 10: 2FA submit button found")

                self.log("STEP 11: Clicking 2FA submit button...")
                await self.page.click(OTP_SUBMIT_SELECTOR)
                self.log("STEP 11: 2FA submit button clicked")

                self.log("STEP 12: Waiting for 2FA submission to process...")
                await self.page.wait_for_timeout(2000)

            except Exception as exc:
                self.log(f"STEP 7-12 ERROR: 2FA process failed - {exc}")
                self.log("STEP 7-12: Checking if we're already logged in (no 2FA required)...")
                # Check if we're already on the dashboard
                try:
                    await self.page.wait_for_selector(LOGIN_SUCCESS_SELECTOR, timeout=3000)
                    self.log("STEP 7-12: Already logged in - 2FA not required for this broker")
                except Exception:
                    raise RuntimeError(f"2FA failed and login not successful: {exc}") from exc

        # Wait for successful login confirmation
        self.log("STEP 13: Waiting for login success confirmation (#menuTableBody)...")
        try:
            await self.page.wait_for_selector(LOGIN_SUCCESS_SELECTOR, timeout=15000)
            self.log("STEP 13: Login success confirmed - dashboard menu found")
        except Exception as exc:
            current_url = self.page.url
            page_content = await self.page.content()
            self.log(f"STEP 13 ERROR: Login confirmation failed")
            self.log(f"STEP 13 ERROR: Current URL: {current_url}")
            self.log(f"STEP 13 ERROR: Page title: {await self.page.title()}")
            raise RuntimeError(f"Login failed - dashboard not found. URL: {current_url}") from exc

        self.log("=" * 60)
        self.log("LOGIN PROCESS COMPLETED SUCCESSFULLY")
        self.log("=" * 60)

    # ------------------------------------------------------------------
    # Session Management
    # ------------------------------------------------------------------
    async def save_session_state(self) -> Dict[str, Any]:
        """
        Save current browser context state (cookies + localStorage).
        Extracts cookie expiry and calculates hint (expiry - 5 minutes).
        
        Returns:
            Dictionary containing storage state with _calculated_expiry hint
        """
        if not self.context:
            raise RuntimeError("Context not initialized")
        
        self.log("Saving browser session state (cookies + storage)...")
        state = await self.context.storage_state()
        
        # Extract earliest cookie expiry from actual cookies
        earliest_expiry = None
        cookies = state.get("cookies", [])
        if cookies:
            # Get all cookie expiry timestamps, filtering out invalid values
            # Session cookies have expires: -1 or very old dates (before 1970)
            # We only want valid future expiry dates
            from datetime import datetime, timezone
            now_timestamp = datetime.now(timezone.utc).timestamp()
            
            expiry_timestamps = []
            for cookie in cookies:
                expires = cookie.get("expires")
                if expires is not None:
                    # Filter out invalid expiry values:
                    # - Negative values (session cookies: -1)
                    # - Very old dates (before 1970, which is timestamp 0)
                    # - Only accept future dates (expires > now)
                    if expires > 0 and expires > now_timestamp:
                        expiry_timestamps.append(expires)
            
            if expiry_timestamps:
                earliest_expiry = min(expiry_timestamps)  # Unix timestamp (seconds)
                # Convert to datetime for logging
                expiry_dt = datetime.fromtimestamp(earliest_expiry, tz=timezone.utc)
                self.log(f"Earliest cookie expires at: {expiry_dt}")
            else:
                self.log("No cookies with valid expiry found (session cookies only - no expiry date)")
        
        # Store the calculated expiry in the state for convenience
        state["_calculated_expiry"] = earliest_expiry
        
        self.log("Session state saved successfully")
        return state

    async def load_session_state(self, state: Dict[str, Any]) -> None:
        """
        Load saved browser context state (cookies + localStorage).
        
        Args:
            state: Dictionary from save_session_state() containing cookies and origins
        """
        if not self.context:
            raise RuntimeError("Context not initialized")
        
        self.log("Loading saved browser session state...")
        
        # Load cookies
        cookies = state.get("cookies", [])
        if cookies:
            await self.context.add_cookies(cookies)
            self.log(f"Loaded {len(cookies)} cookies from saved session")
        else:
            self.log("No cookies found in saved session state")
        
        # Note: localStorage/IndexedDB restoration would require navigating to the domain first
        # For now, cookies are sufficient for NetCHB session persistence
        
        self.log("Session state loaded successfully")

    async def is_session_valid(self) -> bool:
        """
        Check if current session is still valid by accessing a protected page (AMS).
        Uses HTTP request for faster validation (falls back to browser if needed).
        Returns True if we're still logged in, False if redirected to login.
        
        This validates session works for actual operations, not just cookie existence.
        """
        import time
        validation_start = time.time()
        
        self.log("Validating session by accessing protected page (AMS)...")
        
        # Try HTTP validation first (much faster - ~1-2s vs ~5-7s)
        try:
            if self.context:
                storage_state = await self.context.storage_state()
                session_cookies = self._load_cookies_from_storage_state(storage_state)
                
                if session_cookies:
                    self.log("Using HTTP method for session validation (faster)...")
                    http_start = time.time()
                    async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
                        # Set cookies
                        for name, value in session_cookies.items():
                            client.cookies.set(name, value, domain=".netchb.com")
                        
                        # Try to access AMS search page
                        response = await client.get(
                            AMS_SEARCH_URL,
                            headers={
                                "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 18_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.5 Mobile/15E148 Safari/604.1",
                                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                            },
                        )
                        
                        http_time = time.time() - http_start
                        self.log(f"HTTP request completed in {http_time:.2f}s")
                        
                        # Check if we're redirected to login page
                        if response.status_code == 200:
                            # Parse HTML to check if we're on login page or AMS page
                            soup = BeautifulSoup(response.text, "html.parser")
                            
                            # Check for login page indicator
                            login_indicator = soup.find(id="lName")
                            if login_indicator:
                                self.log("Session is invalid - redirected to login page (HTTP check)")
                                validation_time = time.time() - validation_start
                                self.log(f"❌ Session validation failed in {validation_time:.2f}s")
                                return False
                            
                            # Check for AMS page indicator
                            ams_indicator = soup.find(id="pre")
                            if ams_indicator:
                                self.log("Session is valid - can access AMS page (HTTP check)")
                                validation_time = time.time() - validation_start
                                self.log(f"✅ Session validation complete in {validation_time:.2f}s")
                                return True
                            
                            # Check URL
                            final_url = str(response.url)
                            if "security" in final_url.lower() or "login" in final_url.lower():
                                self.log("Session is invalid - on login page (URL check)")
                                validation_time = time.time() - validation_start
                                self.log(f"❌ Session validation failed in {validation_time:.2f}s")
                                return False
                            else:
                                self.log("Session appears valid - not on login page (HTTP check)")
                                validation_time = time.time() - validation_start
                                self.log(f"✅ Session validation complete in {validation_time:.2f}s")
                                return True
        except Exception as exc:
            self.log(f"HTTP validation failed: {exc} - falling back to browser validation")
        
        # Fallback to browser validation (kept since session validation is infrequent)
        if not self.page:
            raise RuntimeError("Page not initialized")
        
        try:
            # Navigate to AMS search page (protected, requires login)
            self.log(f"Using browser method for session validation...")
            self.log(f"Navigating to AMS page: {AMS_SEARCH_URL}")
            nav_start = time.time()
            await self.page.goto(AMS_SEARCH_URL, wait_until="domcontentloaded", timeout=30000)
            self.log(f"Page loaded in {time.time() - nav_start:.2f}s. Current URL: {self.page.url}")
            
            # Wait a bit for any redirects
            await self.page.wait_for_timeout(2000)
            
            # Check if we're on login page (session invalid) or AMS page (session valid)
            try:
                # Check for login page indicator
                await self.page.wait_for_selector("#lName", timeout=5000, state="visible")
                self.log("Session is invalid - redirected to login page")
                validation_time = time.time() - validation_start
                self.log(f"❌ Session validation failed in {validation_time:.2f}s")
                return False
            except Exception:
                # Not on login page, check if we're on AMS page
                try:
                    # Check for AMS page elements (prefix field indicates we're logged in)
                    await self.page.wait_for_selector("#pre", timeout=5000, state="visible")
                    self.log("Session is valid - can access AMS page")
                    validation_time = time.time() - validation_start
                    self.log(f"✅ Session validation complete in {validation_time:.2f}s")
                    return True
                except Exception:
                    # Ambiguous - check URL
                    current_url = self.page.url
                    if "security" in current_url.lower() or "login" in current_url.lower():
                        self.log("Session is invalid - on login page (URL check)")
                        validation_time = time.time() - validation_start
                        self.log(f"❌ Session validation failed in {validation_time:.2f}s")
                        return False
                    else:
                        # Might be on AMS or another protected page
                        self.log("Session appears valid - not on login page")
                        validation_time = time.time() - validation_start
                        self.log(f"✅ Session validation complete in {validation_time:.2f}s")
                        return True
        except Exception as exc:
            self.log(f"Error validating session: {exc}")
            validation_time = time.time() - validation_start
            self.log(f"❌ Session validation failed in {validation_time:.2f}s")
            # On error, assume invalid to be safe
            return False

    # ------------------------------------------------------------------
    # Processing
    # ------------------------------------------------------------------
    async def process_mawb(
        self,
        mawb: str,
        *,
        sections: Dict[str, bool],
        format_identifier: str,
        format_record: Optional[Dict[str, Any]] = None,  # Full format record for template_payload access
        checkbook_hawbs: Optional[str] = None,  # Checkbook HAWBs from batch item (for verification)
        airport_code: Optional[str] = None,  # Airport code for file naming
        customer: Optional[str] = None,  # Customer name for file naming
    ) -> DutyRunResult:
        if not self.page:
            raise RuntimeError("Page not initialized")

        self._logs = []
        digits = _normalize_mawb(mawb)
        result = DutyRunResult(mawb=digits)
        self.log("=" * 60)
        self.log(f"PROCESSING MAWB: {digits}")
        self.log("=" * 60)

        summary: Dict[str, str] = {
            "MAWB Number": digits,
            "AMS Total HAWBs": "N/A",
            "AMS Duty": "N/A",
            "AMS Total T-11 Entries": "N/A",
            "AMS Entries Accepted": "N/A",
            "Rejected Entries": "N/A",
            "7501 Total T-11 Entries": "N/A",
            "7501 Total Houses": "N/A",
            "7501 Duty": "N/A",
            "Report Duty": "N/A",
            "Report Total House": "N/A",
            "Total Informal Duty": "N/A",
            "Complete Total Duty": "N/A",
            "Entry Date": "N/A",
            "Cargo Release Date": "N/A",
            "7501 Batch PDF URL": "N/A",
            "Checkbook HAWBs": str(checkbook_hawbs).strip() if checkbook_hawbs is not None else "N/A",
        }
        
        # Log checkbook_hawbs if provided
        if checkbook_hawbs is not None:
            self.log(f"Checkbook HAWBs provided: '{checkbook_hawbs}' → Summary: '{summary['Checkbook HAWBs']}'")
        else:
            self.log("Checkbook HAWBs not provided (will show as N/A)")

        if sections.get("ams"):
            try:
                self.log("--- Starting AMS Section ---")
                await self._process_ams_section(digits, summary)
                self.log("--- AMS Section Complete ---")
            except Exception as exc:
                self.log(f"⚠️ AMS section failed (skipping - will show N/A values): {exc}")
                import traceback
                self.log(f"AMS section traceback: {traceback.format_exc()}")
                # Keep N/A values in summary (already set above)
        
        # Check if master not found - skip further processing if so
        # Only check if AMS section was enabled (Master Status is only set by AMS section)
        if sections.get("ams") and summary.get("Master Status") == "Not Found":
            self.log("Master not found - skipping entries, custom report, and PDF download sections")
            result.status = "failed"
            result.error_message = "Master not found"
            result.summary = summary
            result.logs = self.logs
            return result

        entries_data = None
        # Entries section runs if:
        # 1. Explicitly enabled (entries=True)
        # 2. Custom report enabled (needs oldest_entry date)
        # 3. PDF download enabled (needs entry_rows for PDF download)
        if sections.get("entries") or sections.get("custom") or sections.get("download_7501_pdf"):
            try:
                self.log("--- Starting Entries Section ---")
                entries_data = await self._process_entries_section(
                    digits, summary
                )
                self.log("--- Entries Section Complete ---")
            except Exception as exc:
                self.log(f"⚠️ Entries section failed (skipping - will show N/A values): {exc}")
                import traceback
                self.log(f"Entries section traceback: {traceback.format_exc()}")
                # Keep N/A values in summary (already set above)
                entries_data = None  # Ensure entries_data is None on failure

        if sections.get("custom"):
            try:
                self.log("--- Starting Custom Report Section ---")
                # Check if entries not found
                if entries_data and entries_data.get("entries_not_found"):
                    self.log("⚠️ Custom report skipped: Entries not found")
                else:
                    oldest_entry = entries_data["oldest_entry"] if entries_data else None
                    if not oldest_entry:
                        self.log("⚠️ Custom report skipped: No oldest entry date available (Entries section may have failed)")
                    else:
                        artifact_path, report_summary = await self._process_custom_report(
                            digits, oldest_entry, format_identifier, format_record
                        )
                        if artifact_path:
                            result.download_path = artifact_path
                            self.log(f"Custom report downloaded to: {artifact_path}")
                        summary.update(report_summary)
                        self.log("--- Custom Report Section Complete ---")
            except Exception as exc:
                self.log(f"⚠️ Custom report failed (skipping - will show N/A values): {exc}")
                import traceback
                self.log(f"Custom report traceback: {traceback.format_exc()}")
                # Keep N/A values in summary (already set above)

        # Handle PDF download if enabled
        if sections.get("download_7501_pdf"):
            try:
                self.log("--- Starting PDF Download Section ---")
                
                # Check if entries not found - skip PDF download if so
                if entries_data and entries_data.get("entries_not_found"):
                    self.log("⚠️ PDF download skipped: Entries not found")
                else:
                    # Get entry_rows from entries_data (reused from entries section)
                    entry_rows = entries_data.get("entry_rows") if entries_data else None
                    
                    if not entry_rows:
                        self.log("⚠️ PDF download skipped: No entry_rows available (Entries section may have failed)")
                    else:
                        # Get session cookies from browser context
                        storage_state = await self.context.storage_state()
                        session_cookies = self._load_cookies_from_storage_state(storage_state)
                        
                        if not session_cookies:
                            self.log("⚠️ PDF download skipped: No session cookies available")
                        else:
                            # Initial verification BEFORE PDF download (only if AMS and Custom Report enabled)
                            should_download = True
                            if sections.get("ams") and sections.get("custom"):
                                try:
                                    def parse_value(val):
                                        if val is None:
                                            return 0
                                        if isinstance(val, (int, float)):
                                            return float(val)
                                        val_str = str(val).replace("$", "").replace(",", "").strip()
                                        try:
                                            return float(val_str)
                                        except (ValueError, AttributeError):
                                            return 0
                                    
                                    ams_hawbs = parse_value(summary.get("AMS Total HAWBs"))
                                    houses_7501 = parse_value(summary.get("7501 Total Houses"))
                                    report_houses = parse_value(summary.get("Report Total House"))
                                    checkbook_hawbs = parse_value(summary.get("Checkbook HAWBs"))
                                    rejected = parse_value(summary.get("Rejected Entries"))
                                    ams_duty = parse_value(summary.get("AMS Duty"))
                                    report_duty = parse_value(summary.get("Report Duty"))
                                    
                                    tolerance = 0.01
                                    
                                    # Initial verification checks (before PDF download)
                                    houses_match = ams_hawbs == houses_7501 == report_houses == checkbook_hawbs
                                    rejected_ok = rejected == 0
                                    duties_match = abs(ams_duty - report_duty) <= tolerance
                                    
                                    if houses_match and rejected_ok and duties_match:
                                        self.log("PDF INITIAL VERIFICATION: ✅ Passed (Houses match, No rejected entries, AMS Duty == Report Duty)")
                                        should_download = True
                                    else:
                                        issues = []
                                        if not houses_match:
                                            issues.append(f"Houses mismatch (AMS: {ams_hawbs}, 7501: {houses_7501}, Report: {report_houses}, Checkbook: {checkbook_hawbs})")
                                        if not rejected_ok:
                                            issues.append(f"Rejected entries: {rejected}")
                                        if not duties_match:
                                            issues.append(f"Duty mismatch (AMS: ${ams_duty:.2f}, Report: ${report_duty:.2f})")
                                        self.log(f"PDF INITIAL VERIFICATION: ❌ Failed - {', '.join(issues)}")
                                        should_download = False
                                except Exception as verify_exc:
                                    self.log(f"PDF INITIAL VERIFICATION ERROR: Failed to verify: {verify_exc}")
                                    # On error, allow download to proceed (fail open)
                                    should_download = True
                            
                            if not should_download:
                                self.log("⚠️ PDF download skipped: Initial verification failed")
                            else:
                                # Download PDF
                                pdf_path = await self._download_7501_batch_pdf_http(
                                    digits,
                                    entry_rows=entry_rows,
                                    session_cookies=session_cookies,
                                )
                                
                                if pdf_path and pdf_path.exists():
                                    # Extract entries and duty from PDF
                                    try:
                                        from .pdf_extractor import extract_entries_and_duty_from_pdf
                                        entry_count, total_duty = extract_entries_and_duty_from_pdf(pdf_path)
                                        
                                        # Always update summary with extracted values (even if 0)
                                        summary["7501 Total T-11 Entries"] = str(entry_count)
                                        summary["7501 Duty"] = f"{total_duty:.2f}"
                                        
                                        if entry_count > 0:
                                            self.log(f"PDF EXTRACTION: Extracted {entry_count} entries from PDF")
                                        else:
                                            self.log("PDF EXTRACTION: ⚠️ Warning - No entries extracted from PDF (count=0)")
                                        
                                        if total_duty > 0:
                                            self.log(f"PDF EXTRACTION: Extracted total duty ${total_duty:.2f} from PDF")
                                        else:
                                            self.log("PDF EXTRACTION: ⚠️ Warning - No duty extracted from PDF (total=0)")
                                    except Exception as extract_exc:
                                        self.log(f"PDF EXTRACTION ERROR: Failed to extract values: {extract_exc}")
                                        import traceback
                                        self.log(f"PDF extraction traceback: {traceback.format_exc()}")
                                        # Keep N/A values if extraction fails
                                    
                                    # Upload PDF to storage
                                    try:
                                        from .storage import NetChbDutyStorageManager
                                        storage = NetChbDutyStorageManager()
                                        
                                        # Use airport_code and customer if provided (for file naming)
                                        pdf_storage_path, pdf_url = storage.upload_pdf(
                                            pdf_path,
                                            digits,
                                            airport_code=airport_code,
                                            customer=customer,
                                        )
                                        
                                        summary["7501 Batch PDF URL"] = pdf_url if pdf_url and pdf_url.strip() else ""
                                        self.log(f"PDF DOWNLOAD: ✓ PDF uploaded to storage: {pdf_storage_path}")
                                        
                                        # Final verification after PDF download (now we have all values)
                                        if sections.get("ams") and sections.get("custom"):
                                            try:
                                                def parse_value(val):
                                                    if val is None:
                                                        return 0
                                                    if isinstance(val, (int, float)):
                                                        return float(val)
                                                    val_str = str(val).replace("$", "").replace(",", "").strip()
                                                    try:
                                                        return float(val_str)
                                                    except (ValueError, AttributeError):
                                                        return 0
                                                
                                                ams_duty = parse_value(summary.get("AMS Duty"))
                                                report_duty = parse_value(summary.get("Report Duty"))
                                                duty_7501 = parse_value(summary.get("7501 Duty"))
                                                ams_t11 = parse_value(summary.get("AMS Total T-11 Entries"))
                                                t11_7501 = parse_value(summary.get("7501 Total T-11 Entries"))
                                                
                                                tolerance = 0.01
                                                
                                                # Final verification checks
                                                duties_match = abs(ams_duty - report_duty) <= tolerance and abs(ams_duty - duty_7501) <= tolerance and abs(report_duty - duty_7501) <= tolerance
                                                t11_match = ams_t11 == t11_7501
                                                
                                                if duties_match and t11_match:
                                                    self.log("PDF FINAL VERIFICATION: ✅ All values match (AMS Duty == Report Duty == 7501 Duty, AMS T-11 == 7501 T-11)")
                                                else:
                                                    issues = []
                                                    if not duties_match:
                                                        issues.append(f"Duty mismatch (AMS: ${ams_duty:.2f}, Report: ${report_duty:.2f}, 7501: ${duty_7501:.2f})")
                                                    if not t11_match:
                                                        issues.append(f"T-11 mismatch (AMS: {ams_t11}, 7501: {t11_7501})")
                                                    self.log(f"PDF FINAL VERIFICATION: ⚠️ Verification issues: {', '.join(issues)}")
                                            except Exception as verify_exc:
                                                self.log(f"PDF FINAL VERIFICATION ERROR: Failed to verify: {verify_exc}")
                                        
                                        # Clean up temp file
                                        try:
                                            pdf_path.unlink()
                                        except Exception:
                                            pass
                                    except Exception as upload_exc:
                                        self.log(f"PDF UPLOAD ERROR: Failed to upload PDF: {upload_exc}")
                                        summary["7501 Batch PDF URL"] = ""
                                else:
                                    self.log("PDF DOWNLOAD: ⚠️ PDF download failed or returned no file")
                
                self.log("--- PDF Download Section Complete ---")
            except Exception as exc:
                self.log(f"⚠️ PDF download section failed: {exc}")
                import traceback
                self.log(f"PDF download traceback: {traceback.format_exc()}")
                # Keep N/A values in summary (already set above)

        result.summary = summary
        result.logs = self.logs
        self.log("=" * 60)
        self.log(f"MAWB PROCESSING COMPLETE: {digits}")
        self.log(f"FINAL SUMMARY CHECK - 7501 Duty: '{summary.get('7501 Duty', 'NOT SET')}'")
        self.log("=" * 60)
        return result

    # ------------------------------------------------------------------
    # Sections
    # ------------------------------------------------------------------
    def _get_session_cookies(self) -> Dict[str, str]:
        """
        Extract cookies from current browser context for HTTP requests.
        
        Returns:
            Dictionary of cookies in format {name: value}
        """
        if not self.context:
            return {}
        
        # Get cookies from current context
        cookies = {}
        # Note: We'll get cookies from storage_state when needed
        # For now, return empty dict - will be populated from storage_state
        return cookies
    
    def _load_cookies_from_storage_state(self, storage_state: Dict) -> Dict[str, str]:
        """
        Convert Playwright storage state cookies to httpx cookie format.
        
        Args:
            storage_state: Dictionary from save_session_state() containing cookies
            
        Returns:
            Dictionary of cookies for httpx client
        """
        cookies = {}
        for cookie in storage_state.get("cookies", []):
            name = cookie.get("name")
            value = cookie.get("value")
            if name and value:
                cookies[name] = value
        return cookies
    
    def _extract_ams_mawb_id(self, html: str, url: Optional[str] = None) -> Optional[str]:
        """
        Extract amsMawbId using multiple methods (fallback chain).
        
        Args:
            html: HTML content
            url: Current URL (optional, for extracting from query params)
            
        Returns:
            amsMawbId string or None if not found
        """
        soup = BeautifulSoup(html, "html.parser")
        
        # Method 1: Extract from master link in search results table (PRIMARY)
        results_div = soup.find("div", id="resultsDiv")
        if results_div:
            table = results_div.find("table")
            if table:
                tbody = table.find("tbody")
                if tbody:
                    rows = tbody.find_all("tr", class_=["light", "dark"])
                    if rows:
                        first_row = rows[0]
                        mawb_link = first_row.find("td")
                        if mawb_link:
                            mawb_link = mawb_link.find("a")
                            if mawb_link:
                                href = mawb_link.get("href", "")
                                if href and "amsMawbId=" in href:
                                    ams_mawb_id = href.split("amsMawbId=")[1].split("&")[0]
                                    self.log(f"  ✓ Extracted amsMawbId from search results link: {ams_mawb_id}")
                                    return ams_mawb_id
        
        # Method 2: Extract from URL query parameter (if already on master page)
        if url and "amsMawbId=" in url:
            ams_mawb_id = url.split("amsMawbId=")[1].split("&")[0]
            self.log(f"  ✓ Extracted amsMawbId from URL: {ams_mawb_id}")
            return ams_mawb_id
        
        self.log("  ⚠ Could not extract amsMawbId using any method")
        return None
    
    def _parse_ams_search_results(self, html: str) -> Optional[Dict]:
        """
        Parse AMS search results HTML and extract data from first result row.
        
        Args:
            html: HTML response from AMS search
            
        Returns:
            Dictionary with extracted data or None if parsing fails
        """
        soup = BeautifulSoup(html, "html.parser")
        
        # Find results table
        results_div = soup.find("div", id="resultsDiv")
        if not results_div:
            self.log("ERROR: resultsDiv not found in HTML")
            return None
        
        table = results_div.find("table")
        if not table:
            self.log("ERROR: Results table not found")
            return None
        
        tbody = table.find("tbody")
        if not tbody:
            self.log("ERROR: tbody not found")
            return None
        
        # Check for "There is no awb" message in the HTML
        page_text = soup.get_text().lower()
        if "there is no awb" in page_text or "no awb" in page_text:
            self.log("Master not found: 'There is no awb' message detected")
            return {"master_not_found": True}
        
        # Find first data row (skip header row)
        rows = tbody.find_all("tr", class_=["light", "dark"])
        if not rows:
            self.log("WARNING: No result rows found")
            return {"master_not_found": True}
        
        first_row = rows[0]
        cells = first_row.find_all("td")
        
        if len(cells) < 7:
            self.log(f"ERROR: Expected at least 7 cells, found {len(cells)}")
            return None
        
        # Extract amsMawbId using multiple methods
        ams_mawb_id = self._extract_ams_mawb_id(html)
        
        # Extract master link from cell 0 (td:nth-child(1))
        mawb_cell = cells[0]
        mawb_link = mawb_cell.find("a")
        master_link = None
        
        if mawb_link:
            href = mawb_link.get("href", "")
            if href:
                # Convert to absolute URL
                if href.startswith("/"):
                    master_link = urljoin("https://www.netchb.com", href)
                else:
                    master_link = href
                # If we didn't get amsMawbId from extract function, try from href
                if not ams_mawb_id and "amsMawbId=" in href:
                    ams_mawb_id = href.split("amsMawbId=")[1].split("&")[0]
        
        # Cell 5 (td:nth-child(6)): Arrival Date - matching Playwright: td:nth-child(6)
        arrival_date = "N/A"
        if len(cells) > 5:
            arrival_cell = cells[5]
            arrival_date = arrival_cell.get_text(strip=True) or "N/A"
        
        # Cell 6 (td:nth-child(7)): Total HAWBs - matching Playwright: td:nth-child(7)
        total_hawbs = "N/A"
        if len(cells) > 6:
            hawbs_cell = cells[6]
            total_hawbs = hawbs_cell.get_text(strip=True) or "N/A"
        
        return {
            "ams_mawb_id": ams_mawb_id,
            "master_link": master_link,
            "total_hawbs": total_hawbs,
            "arrival_date": arrival_date,
        }
    
    def _parse_entries_search_results(self, html: str) -> Optional[Dict]:
        """
        Parse Entries search results HTML and extract entry data.
        
        Args:
            html: HTML response from Entries search
            
        Returns:
            Dictionary with:
            - entry_rows: List of entry data (date, link, query_string)
            - total_entries: Total number of entries found
            - oldest_entry_date: Oldest entry date (datetime object)
        """
        soup = BeautifulSoup(html, "html.parser")
        
        # Find the full table (not just tbody) to access header row
        full_table = soup.select_one("#veForm > div.dataCell > table")
        if not full_table:
            full_table = soup.select_one("div.dataCell > table")
            if full_table:
                self.log("  Using fallback selector: div.dataCell > table")
        
        if not full_table:
            self.log("ERROR: Results table not found (tried both #veForm > div.dataCell > table and div.dataCell > table)")
            return None
        
        # Find tbody for data rows
        results_table = full_table.find("tbody")
        if not results_table:
            self.log("ERROR: Table tbody not found")
            return None
        
        # Find header row to locate "Entry Date" column dynamically
        # Different brokers have Entry Date in different column positions
        # Header selector: #veForm > div.dataCell > table > tbody > tr:nth-child(2) (for most brokers including Allied)
        entry_date_column_idx = None
        
        def search_header_row_for_entry_date(header_row, row_label: str = ""):
            """Helper function to search a header row for Entry Date column."""
            header_cells = header_row.find_all("td")
            for col_idx, header_cell in enumerate(header_cells):
                # Get text from header cell (including nested divs)
                header_text = header_cell.get_text(strip=True)
                # Also check nested div elements (some headers use divs like <div id="eDte_ob">Entry Date</div>)
                divs = header_cell.find_all("div")
                for div in divs:
                    div_text = div.get_text(strip=True)
                    if div_text:
                        header_text = div_text
                        break
                
                # Check for "Entry Date" (case-insensitive, flexible matching)
                header_text_lower = header_text.lower()
                if "entry date" in header_text_lower or "entrydate" in header_text_lower.replace(" ", ""):
                    self.log(f"  ✓ Found 'Entry Date' header in column {col_idx + 1} (0-indexed: {col_idx}, text: '{header_text}'{row_label})")
                    return col_idx
            return None
        
        # Method 1: Try using the specific header row selector (tr:nth-child(2) - second row in tbody)
        # User specified: #veForm > div.dataCell > table > tbody > tr:nth-child(2) for most brokers including Allied
        if results_table:
            tbody_rows = results_table.find_all("tr")
            # Check tr:nth-child(2) (index 1, 0-indexed) - second row
            if len(tbody_rows) > 1:
                header_row_candidate = tbody_rows[1]  # Second row (tr:nth-child(2))
                entry_date_column_idx = search_header_row_for_entry_date(header_row_candidate, ", from tr:nth-child(2)")
            
            # Also check tr:nth-child(1) (first row) as fallback in case structure differs
            if entry_date_column_idx is None and len(tbody_rows) > 0:
                first_row = tbody_rows[0]  # First row (tr:nth-child(1))
                entry_date_column_idx = search_header_row_for_entry_date(first_row, ", from tr:nth-child(1)")
        
        # Method 2: Fallback to finding header rows by class="header"
        if entry_date_column_idx is None:
            header_rows = full_table.find_all("tr", class_="header")
            if header_rows:
                # Search through header rows for "Entry Date"
                # Header rows may have rowspan attributes, but column index is still correct
                for header_row in header_rows:
                    entry_date_column_idx = search_header_row_for_entry_date(header_row, ", from class='header'")
                    if entry_date_column_idx is not None:
                        break
        
        # Fallback: if header not found, use common column indices [5, 6, 4]
        if entry_date_column_idx is None:
            self.log("  ⚠️ Warning: Could not find 'Entry Date' header, will try common column positions [6, 7, 5]")
            # We'll try multiple columns as fallback
        
        rows = results_table.find_all("tr", class_=["light", "dark"])
        if not rows:
            self.log("Entries not found: No entry rows found in results table")
            return {
                "entry_rows": [],
                "entries_not_found": True,
                "total_entries": 0,
                "oldest_entry_date": None,
            }
        
        # Check if first row is a "No Results" message
        first_row_text = rows[0].get_text(strip=True).lower()
        if "no results" in first_row_text or "no entries" in first_row_text:
            self.log(f"Entries not found: No entries found for this MAWB (message: '{rows[0].get_text(strip=True)}')")
            return {
                "entry_rows": [],
                "total_entries": 0,
                "oldest_entry_date": None,
                "entries_not_found": True,
            }
        
        entry_rows = []
        entry_dates = []
        
        for idx, row in enumerate(rows):
            cells = row.find_all("td")
            if len(cells) < 7:
                self.log(f"WARNING: Row {idx+1} has {len(cells)} cells (expected at least 7)")
                # Try to extract what we can even with fewer cells (matching test script)
                if len(cells) >= 1:
                    # Still try to find entry link in first cell
                    link_cell = cells[0]
                    link_elem = link_cell.find("a")
                    if link_elem:
                        href = link_elem.get("href", "")
                        if href:
                            if href.startswith("/"):
                                entry_link = urljoin("https://www.netchb.com", href)
                            else:
                                entry_link = href
                            match = re.search(r"filerCode=[^&]+&entryNo=\d+", entry_link)
                            query_string = match.group(0) if match else None
                            entry_rows.append({
                                "date": None,
                                "date_text": None,
                                "link": entry_link,
                                "query_string": query_string,
                            })
                            self.log(f"  ✓ Extracted entry link from row with {len(cells)} cells: {entry_link}")
                continue
            
            # Extract Entry Date from the column identified in header (or fallback to common positions)
            date_cell = None
            date_text = None
            
            # Try using the column index found from header first
            if entry_date_column_idx is not None:
                if len(cells) > entry_date_column_idx:
                    cell_text = cells[entry_date_column_idx].get_text(strip=True)
                    # Check if it looks like a date (MM/DD/YY format)
                    if cell_text and "/" in cell_text and len(cell_text) <= 10:
                        try:
                            # Try to parse as date
                            datetime.strptime(cell_text, "%m/%d/%y")
                            date_cell = cells[entry_date_column_idx]
                            date_text = cell_text
                            if idx == 0:  # Log only for first row to avoid spam
                                self.log(f"  ✓ Found Entry Date in column {entry_date_column_idx + 1} (from header): {date_text}")
                        except ValueError:
                            pass  # Not a valid date format, try fallback
            
            # Fallback: if header-based lookup failed or header not found, try common column positions
            if not date_text:
                # Try common positions: 5 (column 6), 6 (column 7), 4 (column 5)
                for col_idx in [5, 6, 4]:
                    if len(cells) > col_idx:
                        cell_text = cells[col_idx].get_text(strip=True)
                        # Check if it looks like a date (MM/DD/YY format)
                        if cell_text and "/" in cell_text and len(cell_text) <= 10:
                            try:
                                # Try to parse as date
                                datetime.strptime(cell_text, "%m/%d/%y")
                                date_cell = cells[col_idx]
                                date_text = cell_text
                                if idx == 0 and entry_date_column_idx is None:  # Log only for first row
                                    self.log(f"  ✓ Found Entry Date in column {col_idx + 1} (fallback): {date_text}")
                                break
                            except ValueError:
                                continue
            
            entry_date = None
            if date_text:
                try:
                    entry_date = datetime.strptime(date_text, "%m/%d/%y")
                    entry_dates.append(entry_date)
                except ValueError as exc:
                    self.log(f"WARNING: Failed to parse date '{date_text}': {exc}")
            
            # Cell 1 (td:nth-child(1)): Entry link
            link_cell = cells[0] if len(cells) > 0 else None
            entry_link = None
            query_string = None
            
            if link_cell:
                link_elem = link_cell.find("a")
                if link_elem:
                    href = link_elem.get("href", "")
                    if href:
                        # Convert to absolute URL
                        if href.startswith("/"):
                            entry_link = urljoin("https://www.netchb.com", href)
                        else:
                            entry_link = href
                        
                        # Extract query string (filerCode=...&entryNo=...)
                        match = re.search(r"filerCode=[^&]+&entryNo=\d+", entry_link)
                        if match:
                            query_string = match.group(0)
            
            entry_rows.append({
                "date": entry_date,
                "date_text": date_text if date_cell else None,
                "link": entry_link,
                "query_string": query_string,
            })
        
        oldest_entry_date = min(entry_dates) if entry_dates else None
        
        self.log(f"  Parsed {len(entry_rows)} entry rows, {len(entry_dates)} with dates")
        
        return {
            "entry_rows": entry_rows,
            "total_entries": len(rows),
            "oldest_entry_date": oldest_entry_date,
        }
    
    def _parse_entry_detail_page(self, html: str) -> int:
        """
        Parse entry detail page HTML and count houses.
        
        Args:
            html: HTML response from entry detail page
            
        Returns:
            Number of houses (rows in #invBdy > tr)
        """
        soup = BeautifulSoup(html, "html.parser")
        
        inv_body = soup.find("tbody", id="invBdy")
        if not inv_body:
            return 0
        
        rows = inv_body.find_all("tr")
        return len(rows)
    
    def _parse_print7501_page(self, html: str) -> float:
        """
        Parse print7501 page HTML and extract duty sum.
        
        Args:
            html: HTML response from print7501 page
            
        Returns:
            Sum of duty + fees
        """
        soup = BeautifulSoup(html, "html.parser")
        
        # Find duty table
        duty_table = soup.select_one(
            "#pForm > div:nth-child(1) > div:nth-child(2) > div > div.content > table"
        )
        if not duty_table:
            duty_table = soup.select_one("div.formContainerWithLabel > div.content > table")
        
        if not duty_table:
            return 0.0
        
        tbody = duty_table.find("tbody")
        if tbody:
            rows = tbody.find_all("tr")
        else:
            rows = duty_table.find_all("tr")
        
        if len(rows) < 2:
            return 0.0
        
        # Row 2 (index 1): Duty, Row 4 (index 3): Fees
        duty_text = "0"
        fees_text = "0"
        
        if len(rows) > 1:
            duty_row = rows[1]
            duty_cells = duty_row.find_all("td")
            if len(duty_cells) >= 2:
                if "duty" in duty_cells[0].get_text(strip=True).lower():
                    duty_text = duty_cells[1].get_text(strip=True) or "0"
        
        if len(rows) > 3:
            fees_row = rows[3]
            fees_cells = fees_row.find_all("td")
            if len(fees_cells) >= 2:
                if "fee" in fees_cells[0].get_text(strip=True).lower():
                    fees_text = fees_cells[1].get_text(strip=True) or "0"
        
        def parse_currency(value: str) -> float:
            try:
                cleaned = value.replace(",", "").replace("$", "").strip()
                if not cleaned:
                    return 0.0
                return float(cleaned)
            except Exception:
                return 0.0
        
        duty = parse_currency(duty_text)
        fees = parse_currency(fees_text)
        return duty + fees
    
    def _parse_ams_master_page(self, html: str) -> Dict[str, str]:
        """
        Parse AMS master page HTML and extract duty/entries data.
        
        Args:
            html: HTML response from master page
            
        Returns:
            Dictionary with extracted data
        """
        soup = BeautifulSoup(html, "html.parser")
        
        result = {
            "duty": "N/A",
            "t11_entries": "0",
            "entries_accepted": "0",
            "houses_7501": "0",
        }
        
        # Find #esH (7501 Total Houses) - positioned before #esD
        houses_elem = soup.find(id="esH")
        if houses_elem:
            houses_text = houses_elem.get_text(strip=True)
            # Remove commas from number (e.g., "3,690" -> "3690")
            houses_text_clean = houses_text.replace(",", "").strip()
            try:
                houses_value = int(houses_text_clean) if houses_text_clean else 0
                result["houses_7501"] = str(houses_value)
                self.log(f"AMS HTTP STEP 4: 7501 Houses={houses_value} (from #esH: '{houses_text}')")
            except ValueError as e:
                self.log(f"AMS HTTP STEP 4 ERROR: Failed to parse houses_text '{houses_text}' as int: {e}")
                result["houses_7501"] = "0"
        else:
            self.log("AMS HTTP STEP 4: ⚠️ #esH element not found in HTML")
            result["houses_7501"] = "0"
        
        # Find #esD (AMS Duty) - matching Playwright logic
        duty_elem = soup.find(id="esD")
        if duty_elem:
            result["duty"] = duty_elem.get_text(strip=True) or "N/A"
        
        # Find #esC (Total T-11 Entries) - matching Playwright logic
        t11_elem = soup.find(id="esC")
        if t11_elem:
            t11_text = t11_elem.get_text(strip=True)
            try:
                result["t11_entries"] = str(int(t11_text)) if t11_text else "0"
            except ValueError:
                result["t11_entries"] = "0"
        
        # Find #esA (Entries Accepted) - matching Playwright logic
        accepted_elem = soup.find(id="esA")
        if accepted_elem:
            accepted_text = accepted_elem.get_text(strip=True)
            try:
                result["entries_accepted"] = str(int(accepted_text)) if accepted_text else "0"
            except ValueError:
                result["entries_accepted"] = "0"
        
        return result
    
    async def _process_ams_section_http(self, mawb_digits: str, summary: Dict[str, str]) -> None:
        """
        Process AMS section using HTTP requests (faster than browser automation).
        
        Args:
            mawb_digits: Normalized MAWB (11 digits)
            summary: Dictionary to update with extracted data
        """
        if not self.context:
            raise RuntimeError("Context not initialized")
        
        prefix, number = mawb_digits[:3], mawb_digits[3:]
        
        # Get cookies from current browser context
        storage_state = await self.context.storage_state()
        session_cookies = self._load_cookies_from_storage_state(storage_state)
        
        if not session_cookies:
            raise RuntimeError("No cookies found in session - HTTP method requires valid session cookies")
        
        self.log(f"AMS HTTP: Using {len(session_cookies)} cookies from session")
        
        headers = {
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 18_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.5 Mobile/15E148 Safari/604.1",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Content-Type": "application/x-www-form-urlencoded",
            "Origin": "https://www.netchb.com",
            "Referer": "https://www.netchb.com/app/ams/viewMawbs.do",
            "Cache-Control": "max-age=0",
            "Connection": "keep-alive",
            "DNT": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        }
        
        # Build form payload
        form_data = {
            "prefix": prefix,
            "mawb": number,
            "refNo": "",
            "hawb": "",
            "arrivalBegin": "",
            "arrivalEnd": "",
            "container": "",
            "cbpStatus": "",
            "acasStatus": "",
            "arrivalAirport": "",
            "carrier": "",
            "flight": "",
            "client": "0",
            "clientName": "",
            "searchByProfile": "true",
            "searchTimePeriod": "Y1",
            "location": "0",  # All Locations
            "user": "",  # All Users
            "noPerPage": "25",
            "cfs": "false",
            "pageNo": "0",
            "orderBy": "amb1",
        }
        
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            # Set cookies
            for name, value in session_cookies.items():
                client.cookies.set(name, value, domain=".netchb.com")
            
            # STEP 1: POST to search endpoint
            self.log("AMS HTTP STEP 1: POST to AMS search endpoint...")
            try:
                response = await client.post(
                    AMS_SEARCH_POST_URL,
                    data=form_data,
                    headers=headers,
                )
                response.raise_for_status()
                self.log(f"AMS HTTP STEP 1: Response status {response.status_code}, length {len(response.text)} bytes")
            except Exception as exc:
                self.log(f"AMS HTTP STEP 1 ERROR: Request failed: {exc}")
                raise RuntimeError(f"AMS HTTP STEP 1 failed: {exc}") from exc
            
            # STEP 2: Parse search results
            self.log("AMS HTTP STEP 2: Parsing search results HTML...")
            search_data = self._parse_ams_search_results(response.text)
            
            # Check if master not found
            if search_data and search_data.get("master_not_found"):
                self.log("Master not found for MAWB")
                summary["Master Status"] = "Not Found"
                return
            
            if not search_data:
                raise RuntimeError("AMS HTTP STEP 2 ERROR: Failed to parse search results")
            
            summary["AMS Total HAWBs"] = search_data.get("total_hawbs", "N/A")
            summary["AMS Arrival Date"] = search_data.get("arrival_date", "N/A")
            self.log(f"AMS HTTP STEP 2: Total HAWBs={summary['AMS Total HAWBs']}, Arrival={summary['AMS Arrival Date']}")
            
            master_link = search_data.get("master_link")
            if not master_link:
                raise RuntimeError("AMS HTTP STEP 2 ERROR: No master link found in search results")
            
            # STEP 3: GET master page
            self.log(f"AMS HTTP STEP 3: GET master page: {master_link}")
            try:
                master_response = await client.get(
                    master_link,
                    headers={
                        "User-Agent": headers["User-Agent"],
                        "Accept": headers["Accept"],
                        "Referer": AMS_SEARCH_POST_URL,
                    },
                )
                master_response.raise_for_status()
                self.log(f"AMS HTTP STEP 3: Master page response status {master_response.status_code}, length {len(master_response.text)} bytes")
            except Exception as exc:
                self.log(f"AMS HTTP STEP 3 ERROR: Master page request failed: {exc}")
                raise RuntimeError(f"AMS HTTP STEP 3 failed: {exc}") from exc
            
            # STEP 4: Parse master page
            self.log("AMS HTTP STEP 4: Parsing master page HTML...")
            master_data = self._parse_ams_master_page(master_response.text)
            
            summary["AMS Duty"] = master_data.get("duty", "N/A")
            summary["AMS Total T-11 Entries"] = master_data.get("t11_entries", "0")
            summary["AMS Entries Accepted"] = master_data.get("entries_accepted", "0")
            summary["7501 Total Houses"] = master_data.get("houses_7501", "N/A")
            
            try:
                t11 = int(master_data.get("t11_entries", "0"))
                accepted = int(master_data.get("entries_accepted", "0"))
                rejected = t11 - accepted
            except (ValueError, TypeError):
                rejected = 0
            
            summary["Rejected Entries"] = str(rejected)
            
            self.log(f"AMS HTTP STEP 4: Duty={summary['AMS Duty']}, T-11={summary['AMS Total T-11 Entries']}, Accepted={summary['AMS Entries Accepted']}, 7501 Houses={summary['7501 Total Houses']}")
            self.log("AMS HTTP: AMS section complete (HTTP method)")

    async def _process_ams_section(self, mawb_digits: str, summary: Dict[str, str]) -> None:
        """
        Process AMS section using HTTP requests only.
        ARCHIVED: Browser method fallback removed to reduce server load.
        
        Args:
            mawb_digits: Normalized MAWB (11 digits)
            summary: Dictionary to update with extracted data
        """
        # HTTP method only - no fallback to reduce server load
        await self._process_ams_section_http(mawb_digits, summary)
    
    async def _process_ams_section_browser(self, mawb_digits: str, summary: Dict[str, str]) -> None:
        """
        ARCHIVED: Process AMS section using browser automation (no longer used - fallback removed).
        
        Args:
            mawb_digits: Normalized MAWB (11 digits)
            summary: Dictionary to update with extracted data
        """
        assert self.page
        prefix, number = mawb_digits[:3], mawb_digits[3:]
        
        self.log("AMS STEP 1: Opening AMS search page in new tab...")
        ams_page = await self.context.new_page()
        await ams_page.goto(AMS_SEARCH_URL, wait_until="domcontentloaded", timeout=30000)
        self.log(f"AMS STEP 1: AMS page loaded. URL: {ams_page.url}")

        self.log("AMS STEP 2: Waiting for prefix field (#pre)...")
        await ams_page.wait_for_selector("#pre", timeout=20000)
        self.log("AMS STEP 2: Prefix field found")

        self.log(f"AMS STEP 3: Filling prefix: {prefix}")
        await ams_page.fill("#pre", prefix)
        self.log("AMS STEP 3: Prefix filled")

        self.log(f"AMS STEP 4: Filling MAWB number: {number}")
        await ams_page.fill("#mawb", number)
        self.log("AMS STEP 4: MAWB number filled")

        self.log("AMS STEP 5: Selecting 'All Locations'...")
        await ams_page.select_option(
            "#mF > div > div.content > table > tbody > tr:nth-child(6) > td:nth-child(1) > select",
            "0"
        )
        self.log("AMS STEP 5: Location selected")

        self.log("AMS STEP 6: Selecting 'All Users'...")
        await ams_page.select_option("#usr", "")
        self.log("AMS STEP 6: User selected")

        self.log("AMS STEP 7: Clicking search button...")
        await ams_page.click("#mF > div > div.content > table > tbody > tr:nth-child(6) > td:nth-child(4) > input[type=submit]")
        self.log("AMS STEP 7: Search button clicked")

        self.log("AMS STEP 8: Waiting for search results (timeout: 60s - site is slow)...")
        await ams_page.wait_for_selector(
            "#resultsDiv > table > tbody > tr.light, #resultsDiv > table > tbody > tr.dark",
            timeout=60000  # 60 seconds - site is very slow
        )
        self.log("AMS STEP 8: Search results loaded")

        rows = await ams_page.query_selector_all(
            "#resultsDiv > table > tbody > tr.light, #resultsDiv > table > tbody > tr.dark"
        )
        self.log(f"AMS STEP 9: Found {len(rows)} result rows")

        # Check for "There is no awb" message on the page
        page_content = await ams_page.content()
        page_text = page_content.lower()
        if "there is no awb" in page_text or "no awb" in page_text:
            self.log("Master not found: 'There is no awb' message detected")
            summary["Master Status"] = "Not Found"
            await ams_page.close()
            return

        if not rows:
            self.log("AMS STEP 9: No results found for this MAWB - Master not found")
            summary["Master Status"] = "Not Found"
            await ams_page.close()
            return
        else:
            hawbs = await rows[0].query_selector("td:nth-child(7)")
            arrival = await rows[0].query_selector("td:nth-child(6)")
            summary["AMS Total HAWBs"] = (await hawbs.inner_text()).strip() if hawbs else "N/A"
            summary["AMS Arrival Date"] = (await arrival.inner_text()).strip() if arrival else "N/A"
            self.log(f"AMS STEP 9: Total HAWBs: {summary['AMS Total HAWBs']}, Arrival: {summary['AMS Arrival Date']}")

            master_link_elem = await rows[0].query_selector("td:nth-child(1) > a")
            if master_link_elem:
                master_link = await master_link_elem.get_attribute("href")
                if not master_link:
                    self.log("AMS STEP 10 ERROR: Master link href is empty")
                else:
                    # Convert relative URL to absolute URL
                    if not master_link.startswith("http"):
                        base_url = ams_page.url
                        master_link = urljoin(base_url, master_link)
                        self.log(f"AMS STEP 10: Converted relative URL to absolute: {master_link}")
                    else:
                        self.log(f"AMS STEP 10: Master link is already absolute: {master_link}")
                    
                    self.log(f"AMS STEP 10: Opening master link (timeout: 60s - site is slow): {master_link}")
                    await ams_page.goto(master_link, wait_until="domcontentloaded", timeout=60000)  # 60 seconds
                    self.log("AMS STEP 10: Master page loaded")

                try:
                    self.log("AMS STEP 11: Waiting for master page elements (timeout: 30s)...")
                    await ams_page.wait_for_selector("#esD", timeout=30000)  # 30 seconds
                    houses_elem = await ams_page.query_selector("#esH")
                    duty_elem = await ams_page.query_selector("#esD")
                    t11_elem = await ams_page.query_selector("#esC")
                    accepted_elem = await ams_page.query_selector("#esA")

                    if houses_elem:
                        houses_text = (await houses_elem.inner_text()).strip()
                        # Remove commas from number (e.g., "3,690" -> "3690")
                        houses_text_clean = houses_text.replace(",", "").strip()
                        try:
                            houses_7501 = int(houses_text_clean) if houses_text_clean else 0
                        except ValueError:
                            self.log(f"AMS STEP 11 ERROR: Failed to parse houses_text '{houses_text}' as int")
                            houses_7501 = 0
                    else:
                        self.log("AMS STEP 11: ⚠️ #esH element not found")
                        houses_7501 = 0
                    duty = (await duty_elem.inner_text()).strip() if duty_elem else "N/A"
                    t11_entries = int((await t11_elem.inner_text()).strip()) if t11_elem else 0
                    entries_accepted = int((await accepted_elem.inner_text()).strip()) if accepted_elem else 0

                    summary["7501 Total Houses"] = str(houses_7501)
                    summary["AMS Duty"] = duty
                    summary["AMS Total T-11 Entries"] = str(t11_entries)
                    summary["AMS Entries Accepted"] = str(entries_accepted)
                    summary["Rejected Entries"] = str(t11_entries - entries_accepted)
                    self.log(f"AMS STEP 11: Duty={duty}, T-11={t11_entries}, Accepted={entries_accepted}, 7501 Houses={houses_7501}")
                except Exception as exc:
                    self.log(f"AMS STEP 11 ERROR: Failed to parse master page: {exc}")

        await ams_page.close()
        self.log("AMS STEP 12: AMS tab closed")

    async def _process_entries_section(
        self,
        mawb_digits: str,
        summary: Dict[str, str],
    ) -> Dict[str, Optional[datetime]]:
        """
        Process Entries section using HTTP requests only.
        ARCHIVED: Browser method fallback removed to reduce server load.
        
        Args:
            mawb_digits: Normalized MAWB (11 digits)
            summary: Dictionary to update with extracted data
            
        Returns:
            Dictionary with oldest_entry date
        """
        # HTTP method only - no fallback to reduce server load
        return await self._process_entries_section_http(mawb_digits, summary)
    
    async def _process_entries_section_http(
        self,
        mawb_digits: str,
        summary: Dict[str, str],
    ) -> Dict[str, Optional[datetime]]:
        """
        Process Entries section using HTTP requests (faster than browser automation).
        
        Args:
            mawb_digits: Normalized MAWB (11 digits)
            summary: Dictionary to update with extracted data
            
        Returns:
            Dictionary with oldest_entry date
        """
        if not self.context:
            raise RuntimeError("Context not initialized")
        
        # Get cookies from current browser context
        storage_state = await self.context.storage_state()
        session_cookies = self._load_cookies_from_storage_state(storage_state)
        
        if not session_cookies:
            raise RuntimeError("No cookies found in session - HTTP method requires valid session cookies")
        
        self.log(f"ENTRIES HTTP: Using {len(session_cookies)} cookies from session")
        
        headers = {
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 18_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.5 Mobile/15E148 Safari/604.1",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Content-Type": "application/x-www-form-urlencoded",
            "Origin": "https://www.netchb.com",
            "Referer": "https://www.netchb.com/app/entry/index.jsp",
            "Cache-Control": "max-age=0",
            "Connection": "keep-alive",
            "DNT": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        }
        
        # Build form payload (matching test script - includes all fields from network capture)
        form_data = {
            "entryNoSearch": "",
            "brokerRefNo": "",
            "importerRecord": "0",
            "importerRecordName": "",
            "importerSearchByProfile": "true",
            "ultimateConsignee": "0",
            "ultimateConsigneeName": "",
            "ultimateConsigneeSearchByProfile": "true",
            "freightForwarder": "0",
            "freightForwarderName": "",
            "freightForwarderSearchByProfile": "true",
            "begin": "",
            "end": "",
            "entryStatus": "",
            "cargoReleaseStatus": "",
            "manifestStatus": "",
            "pgaAgency": "",
            "ogaStatus": "",
            "statusColor": "",
            "entryType": "",
            "portEntry": "",
            "modeTransport": "",
            "masterBill": mawb_digits,  # Full 11-digit MAWB
            "searchTimePeriod": "Y1",  # 1 Year
            "user": "",  # All Users (empty string)
            "location": "0",  # All Locations
            "noPerPage": "1000",  # Show 1000 per page
            "entryNo": "0",
            "orderBy": "vep1",  # Order by Entry No
        }
        
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            # Set cookies
            for name, value in session_cookies.items():
                client.cookies.set(name, value, domain=".netchb.com")
            
            # STEP 1: POST to search endpoint
            self.log("ENTRIES HTTP STEP 1: POST to Entries search endpoint...")
            self.log(f"ENTRIES HTTP STEP 1: Payload: masterBill={mawb_digits}, location=0, user=, noPerPage=1000, searchTimePeriod=Y1")
            try:
                response = await client.post(
                    ENTRIES_SEARCH_POST_URL,
                    data=form_data,
                    headers=headers,
                )
                response.raise_for_status()
                self.log(f"ENTRIES HTTP STEP 1: Response status {response.status_code}, length {len(response.text)} bytes")
                # Debug: Save first 500 chars of response to check if it's the right page
                response_preview = response.text[:500].replace('\n', ' ').replace('\r', ' ')
                self.log(f"ENTRIES HTTP STEP 1: Response preview (first 500 chars): {response_preview}")
            except Exception as exc:
                self.log(f"ENTRIES HTTP STEP 1 ERROR: Request failed: {exc}")
                raise RuntimeError(f"ENTRIES HTTP STEP 1 failed: {exc}") from exc
            
            # STEP 2: Parse search results
            self.log("ENTRIES HTTP STEP 2: Parsing search results HTML...")
            search_data = self._parse_entries_search_results(response.text)
            
            if not search_data:
                raise RuntimeError("ENTRIES HTTP STEP 2 ERROR: Failed to parse search results")
            
            # Check if entries not found
            if search_data.get("entries_not_found"):
                self.log("Entries not found for this MAWB")
                summary["Entries Status"] = "Not Found"
                return {
                    "oldest_entry": None,
                    "entries_not_found": True,
                }
            
            oldest_entry = search_data.get("oldest_entry_date")
            if oldest_entry:
                summary["Entry Date"] = oldest_entry.strftime("%m/%d/%y")
                self.log(f"ENTRIES HTTP STEP 2: Oldest entry date: {summary['Entry Date']}")
            
            # ARCHIVED: STEP 3 - Entry details scraping removed
            # "7501 Total T-11 Entries" now extracted from PDF (Phase 4)
            # "7501 Total Houses" now extracted from AMS section (Phase 2)
            # "7501 Duty" now extracted from PDF (Phase 4)
            # Values remain "N/A" here if PDF download not enabled
            
            self.log("ENTRIES HTTP: Entries section complete (HTTP method)")
            # Store entry_rows for potential reuse in PDF download
            entry_rows = search_data.get("entry_rows", [])
            self._last_entry_rows = entry_rows
            return {
                "oldest_entry": oldest_entry,
                "entry_rows": entry_rows,
                "entries_not_found": search_data.get("entries_not_found", False),
            }
    
    async def _process_entries_section_browser(
        self,
        mawb_digits: str,
        summary: Dict[str, str],
    ) -> Dict[str, Optional[datetime]]:
        """
        ARCHIVED: Process Entries section using browser automation (no longer used - fallback removed).
        
        Args:
            mawb_digits: Normalized MAWB (11 digits)
            summary: Dictionary to update with extracted data
            
        Returns:
            Dictionary with oldest_entry date and entry_rows for PDF download
        """
        assert self.page
        self.log("ENTRIES STEP 1: Opening Entries page in new tab...")
        entries_page = await self.context.new_page()
        await entries_page.goto(ENTRIES_URL, wait_until="domcontentloaded", timeout=30000)
        self.log(f"ENTRIES STEP 1: Entries page loaded. URL: {entries_page.url}")

        self.log("ENTRIES STEP 2: Waiting for MAWB field (#mbl)...")
        await entries_page.wait_for_selector("#mbl", timeout=20000)
        self.log("ENTRIES STEP 2: MAWB field found")

        self.log(f"ENTRIES STEP 3: Filling MAWB: {mawb_digits}")
        await entries_page.fill("#mbl", mawb_digits)
        self.log("ENTRIES STEP 3: MAWB filled")

        self.log("ENTRIES STEP 4: Selecting 'All Users'...")
        await entries_page.select_option("#usr", "")
        self.log("ENTRIES STEP 4: User selected")

        self.log("ENTRIES STEP 5: Selecting 'All Locations'...")
        await entries_page.select_option("#fieldsDiv > table:nth-child(4) > tbody > tr > td:nth-child(3) > select", "0")
        self.log("ENTRIES STEP 5: Location selected")

        self.log("ENTRIES STEP 6: Selecting '1000' for number of pages...")
        await entries_page.select_option("#nop", "1000")
        self.log("ENTRIES STEP 6: Number of pages selected")

        self.log("ENTRIES STEP 7: Clicking search button (#subB)...")
        await entries_page.click("#subB")
        self.log("ENTRIES STEP 7: Search button clicked")

        self.log("ENTRIES STEP 8: Waiting for results table (timeout: 60s - site is slow)...")
        await entries_page.wait_for_selector("#veForm > div.dataCell > table > tbody", timeout=60000)  # 60 seconds
        self.log("ENTRIES STEP 8: Results table loaded")

        # Find Entry Date column index from header row dynamically
        entry_date_column_idx = None
        try:
            # Get the full table HTML to parse headers
            table_html = await entries_page.locator("#veForm > div.dataCell > table").inner_html()
            soup = BeautifulSoup(table_html, "html.parser")
            
            # Find header rows
            header_rows = soup.find_all("tr", class_="header")
            if header_rows:
                # Search through header rows for "Entry Date"
                for header_row in header_rows:
                    header_cells = header_row.find_all("td")
                    for col_idx, header_cell in enumerate(header_cells):
                        # Get text from header cell (including nested divs)
                        header_text = header_cell.get_text(strip=True)
                        # Also check nested div elements
                        divs = header_cell.find_all("div")
                        for div in divs:
                            div_text = div.get_text(strip=True)
                            if div_text:
                                header_text = div_text
                                break
                        
                        # Check for "Entry Date" (case-insensitive)
                        header_text_lower = header_text.lower()
                        if "entry date" in header_text_lower or "entrydate" in header_text_lower.replace(" ", ""):
                            entry_date_column_idx = col_idx
                            self.log(f"ENTRIES STEP 8.5: ✓ Found 'Entry Date' header in column {col_idx + 1} (0-indexed: {col_idx}, text: '{header_text}')")
                            break
                    if entry_date_column_idx is not None:
                        break
            
            if entry_date_column_idx is None:
                self.log("ENTRIES STEP 8.5: ⚠️ Warning: Could not find 'Entry Date' header, will use fallback column 6")
        except Exception as exc:
            self.log(f"ENTRIES STEP 8.5: ⚠️ Warning: Failed to find Entry Date header ({exc}), will use fallback column 6")

        rows = await entries_page.query_selector_all("#veForm > div.dataCell > table > tbody tr.light, #veForm > div.dataCell > table > tbody tr.dark")
        self.log(f"ENTRIES STEP 9: Found {len(rows)} entry rows")

        entry_dates: List[datetime] = []
        entry_links: List[str] = []
        query_strings: List[str] = []

        for idx, row in enumerate(rows):
            try:
                # Use dynamically found column index, or fallback to column 6
                column_to_use = entry_date_column_idx if entry_date_column_idx is not None else 5  # 0-indexed: 5 = column 6
                date_elem = await row.query_selector(f"td:nth-child({column_to_use + 1})")
                if date_elem:
                    date_text = (await date_elem.inner_text()).strip()
                    if date_text:
                        # Validate it looks like a date before parsing
                        if "/" in date_text and len(date_text) <= 10:
                            try:
                                parsed_date = datetime.strptime(date_text, "%m/%d/%y")
                                entry_dates.append(parsed_date)
                                if idx == 0:  # Log only for first row to avoid spam
                                    self.log(f"ENTRIES STEP 9.{idx+1}: Entry date found in column {column_to_use + 1}: {date_text}")
                            except ValueError:
                                # Not a valid date format, try fallback columns
                                if entry_date_column_idx is not None:
                                    # Try common fallback positions
                                    for fallback_col in [5, 6, 4]:  # Columns 6, 7, 5
                                        if fallback_col != column_to_use:
                                            try:
                                                fallback_elem = await row.query_selector(f"td:nth-child({fallback_col + 1})")
                                                if fallback_elem:
                                                    fallback_text = (await fallback_elem.inner_text()).strip()
                                                    if fallback_text and "/" in fallback_text:
                                                        parsed_date = datetime.strptime(fallback_text, "%m/%d/%y")
                                                        entry_dates.append(parsed_date)
                                                        self.log(f"ENTRIES STEP 9.{idx+1}: Entry date found in fallback column {fallback_col + 1}: {fallback_text}")
                                                        break
                                            except:
                                                continue
            except Exception as exc:
                self.log(f"ENTRIES STEP 9.{idx+1}: Failed to parse date: {exc}")
                continue

            # Extract entry links (needed for PDF download in Phase 4)
            try:
                link_elem = await row.query_selector("td:nth-child(1) a")
                if link_elem:
                    link = await link_elem.get_attribute("href")
                    if link:
                        # Convert relative URL to absolute URL
                        if not link.startswith("http"):
                            base_url = entries_page.url
                            link = urljoin(base_url, link)
                            self.log(f"ENTRIES STEP 9.{idx+1}: Converted relative entry URL to absolute: {link}")
                        entry_links.append(link)
                        match = re.search(r"filerCode=[^&]+&entryNo=\d+", link)
                        if match:
                            query_strings.append(match.group(0))
            except Exception as exc:
                self.log(f"ENTRIES STEP 9.{idx+1}: Failed to extract link: {exc}")

        oldest_entry = min(entry_dates) if entry_dates else None
        if oldest_entry:
            summary["Entry Date"] = oldest_entry.strftime("%m/%d/%y")
            self.log(f"ENTRIES STEP 10: Oldest entry date: {summary['Entry Date']}")

        # ARCHIVED: STEP 11 - T-11 entries count removed (now extracted from PDF in Phase 4)
        # ARCHIVED: STEP 12 - Entry details scraping removed
        # "7501 Total T-11 Entries" now extracted from PDF (Phase 4)
        # "7501 Total Houses" now extracted from AMS section (Phase 2)
        # "7501 Duty" now extracted from PDF (Phase 4)
        # Values remain "N/A" here if PDF download not enabled
        
        # Convert entry_links to entry_rows format for PDF download reuse
        entry_rows = [
            {"link": link, "query_string": qs}
            for link, qs in zip(entry_links, query_strings)
        ]

        await entries_page.close()
        self.log("ENTRIES STEP 13: Entries tab closed")
        return {
            "oldest_entry": oldest_entry,
            "entry_rows": entry_rows,
            "entries_not_found": False,
        }

    # ARCHIVED: No longer used - values extracted from PDF instead
    async def _scrape_entries_details_http(
        self,
        client: httpx.AsyncClient,
        session_cookies: Dict[str, str],
        entry_rows: List[Dict],
        headers: Dict[str, str],
    ) -> Tuple[int, float, int, int]:
        """
        ARCHIVED: Scrape entry details using HTTP requests with parallel processing and retry logic.
        No longer used - "7501 Total Houses" and "7501 Duty" are now extracted from PDF.
        
        Args:
            client: httpx client with cookies already set
            session_cookies: Session cookies for setting on new requests
            entry_rows: List of entry row dictionaries with 'link' and 'query_string'
            headers: HTTP headers to use
            
        Returns:
            Tuple of (total_houses, total_duty, entry_detail_failures, print7501_failures)
        """
        entries_total_houses = 0
        entries_total_duty = 0.0
        entry_detail_failures = 0
        print7501_failures = 0
        batch_size = 6  # Process 6 at a time (user requested)
        
        entry_links = [row["link"] for row in entry_rows if row.get("link")]
        query_strings = [row["query_string"] for row in entry_rows if row.get("query_string")]
        
        self.log(f"ENTRIES_DETAILS HTTP STEP 1: Processing {len(entry_links)} entry links in batches of {batch_size}...")
        
        # Process entry detail pages
        async def fetch_entry_houses(link: str, idx: int, total: int) -> Tuple[int, bool]:
            """Fetch and parse a single entry detail page with retry logic."""
            max_retries = 3
            retry_delay = 1.0
            timeout_seconds = 120.0  # 2 minutes for entry detail pages
            
            for attempt in range(max_retries):
                try:
                    async with httpx.AsyncClient(timeout=timeout_seconds, follow_redirects=True) as req_client:
                        for name, value in session_cookies.items():
                            req_client.cookies.set(name, value, domain=".netchb.com")
                        
                        response = await req_client.get(
                            link,
                            headers={
                                "User-Agent": headers["User-Agent"],
                                "Accept": headers["Accept"],
                                "Referer": ENTRIES_SEARCH_POST_URL,
                            },
                        )
                        response.raise_for_status()
                        house_count = self._parse_entry_detail_page(response.text)
                        return (house_count, True)
                except (httpx.TimeoutException, httpx.ConnectError, httpx.NetworkError) as exc:
                    if attempt < max_retries - 1:
                        wait_time = retry_delay * (2 ** attempt)
                        await asyncio.sleep(wait_time)
                        continue
                    else:
                        self.log(f"ENTRIES_DETAILS HTTP: ✗ Entry {idx+1} failed after {max_retries} attempts (network/timeout): {exc}")
                        return (0, False)
                except httpx.HTTPStatusError as exc:
                    if exc.response.status_code >= 500 and attempt < max_retries - 1:
                        wait_time = retry_delay * (2 ** attempt)
                        await asyncio.sleep(wait_time)
                        continue
                    else:
                        self.log(f"ENTRIES_DETAILS HTTP: ✗ Entry {idx+1} failed after {max_retries} attempts (HTTP {exc.response.status_code}): {exc}")
                        return (0, False)
                except Exception as exc:
                    self.log(f"ENTRIES_DETAILS HTTP: ✗ Entry {idx+1} failed: {exc}")
                    return (0, False)
            
            return (0, False)
        
        # Process in batches
        for start in range(0, len(entry_links), batch_size):
            batch = entry_links[start:start + batch_size]
            self.log(f"ENTRIES_DETAILS HTTP STEP 1.{start//batch_size + 1}: Processing batch {start//batch_size + 1} ({len(batch)} entries)...")
            
            tasks = [fetch_entry_houses(link, start + idx, len(entry_links)) for idx, link in enumerate(batch)]
            results = await asyncio.gather(*tasks)
            
            for house_count, success in results:
                entries_total_houses += house_count
                if not success:
                    entry_detail_failures += 1
            
            await asyncio.sleep(0.5)  # Small delay between batches
        
        self.log(f"ENTRIES_DETAILS HTTP STEP 1: Total houses: {entries_total_houses}")
        
        # Process print7501 pages
        self.log(f"ENTRIES_DETAILS HTTP STEP 2: Processing {len(query_strings)} print7501 pages in batches of {batch_size}...")
        
        async def fetch_print7501_duty(query: str, idx: int, total: int) -> Tuple[float, bool]:
            """Fetch and parse a single print7501 page with retry logic."""
            max_retries = 3
            retry_delay = 1.0
            timeout_seconds = 360.0  # 6 minutes for print7501 pages (user requested)
            
            url = f"{PRINT7501_URL}?{query}"
            
            for attempt in range(max_retries):
                try:
                    async with httpx.AsyncClient(timeout=timeout_seconds, follow_redirects=True) as req_client:
                        for name, value in session_cookies.items():
                            req_client.cookies.set(name, value, domain=".netchb.com")
                        
                        response = await req_client.get(
                            url,
                            headers={
                                "User-Agent": headers["User-Agent"],
                                "Accept": headers["Accept"],
                                "Referer": ENTRY_DETAIL_URL,
                            },
                        )
                        response.raise_for_status()
                        duty_sum = self._parse_print7501_page(response.text)
                        return (duty_sum, True)
                except (httpx.TimeoutException, httpx.ConnectError, httpx.NetworkError) as exc:
                    if attempt < max_retries - 1:
                        wait_time = retry_delay * (2 ** attempt)
                        await asyncio.sleep(wait_time)
                        continue
                    else:
                        self.log(f"ENTRIES_DETAILS HTTP: ✗ Print7501 {idx+1} failed after {max_retries} attempts (network/timeout): {exc}")
                        return (0.0, False)
                except httpx.HTTPStatusError as exc:
                    if exc.response.status_code >= 500 and attempt < max_retries - 1:
                        wait_time = retry_delay * (2 ** attempt)
                        await asyncio.sleep(wait_time)
                        continue
                    else:
                        self.log(f"ENTRIES_DETAILS HTTP: ✗ Print7501 {idx+1} failed after {max_retries} attempts (HTTP {exc.response.status_code}): {exc}")
                        return (0.0, False)
                except Exception as exc:
                    self.log(f"ENTRIES_DETAILS HTTP: ✗ Print7501 {idx+1} failed: {exc}")
                    return (0.0, False)
            
            return (0.0, False)
        
        # Process in batches
        for start in range(0, len(query_strings), batch_size):
            batch = query_strings[start:start + batch_size]
            self.log(f"ENTRIES_DETAILS HTTP STEP 2.{start//batch_size + 1}: Processing batch {start//batch_size + 1} ({len(batch)} queries)...")
            
            tasks = [fetch_print7501_duty(query, start + idx, len(query_strings)) for idx, query in enumerate(batch)]
            results = await asyncio.gather(*tasks)
            
            for duty_sum, success in results:
                entries_total_duty += duty_sum
                if not success:
                    print7501_failures += 1
            
            await asyncio.sleep(0.5)  # Small delay between batches
        
        self.log(f"ENTRIES_DETAILS HTTP STEP 2: Total duty: {entries_total_duty:.2f}")
        return entries_total_houses, entries_total_duty, entry_detail_failures, print7501_failures
    
    # ARCHIVED: No longer used - values extracted from PDF instead
    async def _scrape_entries_details(self, entry_links: List[str], query_strings: List[str]) -> Tuple[int, float]:
        assert self.context
        entries_total_houses = 0
        entries_total_duty = 0.0
        group_size = 6  # Updated to 6 (user requested)

        self.log(f"ENTRIES_DETAILS STEP 1: Processing {len(entry_links)} entry links in groups of {group_size}...")

        # Process entry detail pages
        for start in range(0, len(entry_links), group_size):
            group = entry_links[start : start + group_size]
            self.log(f"ENTRIES_DETAILS STEP 1.{start//group_size + 1}: Processing group {start//group_size + 1} ({len(group)} entries)...")
            
            tasks = []
            for link in group:
                tasks.append(self._scrape_single_entry(link))
            
            results = await asyncio.gather(*tasks, return_exceptions=True)
            for idx, result in enumerate(results):
                if isinstance(result, Exception):
                    self.log(f"ENTRIES_DETAILS: ⚠️ Entry {idx+1} in group failed (skipped): {result}")
                else:
                    entries_total_houses += result
                    self.log(f"ENTRIES_DETAILS: Entry {idx+1} in group: {result} houses")

        self.log(f"ENTRIES_DETAILS STEP 1: Total houses: {entries_total_houses}")

        # Process print7501 pages
        print_url = "https://www.netchb.com/app/entry/print7501.do?"
        self.log(f"ENTRIES_DETAILS STEP 2: Processing {len(query_strings)} print7501 pages in groups of {group_size}...")

        for start in range(0, len(query_strings), group_size):
            group = query_strings[start : start + group_size]
            self.log(f"ENTRIES_DETAILS STEP 2.{start//group_size + 1}: Processing group {start//group_size + 1} ({len(group)} queries)...")
            
            tasks = []
            for query in group:
                tasks.append(self._scrape_single_print7501(f"{print_url}{query}"))
            
            results = await asyncio.gather(*tasks, return_exceptions=True)
            for idx, result in enumerate(results):
                if isinstance(result, Exception):
                    self.log(f"ENTRIES_DETAILS: ⚠️ Print7501 {idx+1} in group failed (skipped): {result}")
                else:
                    entries_total_duty += result
                    self.log(f"ENTRIES_DETAILS: Print7501 {idx+1} in group: ${result:.2f} duty")

        self.log(f"ENTRIES_DETAILS STEP 2: Total duty: {entries_total_duty:.2f}")
        return entries_total_houses, entries_total_duty

    # ARCHIVED: No longer used - values extracted from PDF instead
    async def _scrape_single_entry(self, link: str) -> int:
        """
        Scrape a single entry detail page and return house count.
        NOTE: NetCHB site is very slow - using extended timeouts.
        If page fails to load, returns 0 (skips this entry gracefully).
        """
        assert self.context
        page = await self.context.new_page()
        try:
            self.log(f"ENTRIES_DETAILS: Loading entry page (timeout: 120s): {link}")
            await page.goto(link, wait_until="domcontentloaded", timeout=120000)  # 2 minutes for slow site
            self.log(f"ENTRIES_DETAILS: Waiting for #invBdy selector (timeout: 120s)...")
            await page.wait_for_selector("#invBdy", timeout=120000, state="attached")  # 2 minutes for slow site
            rows = await page.query_selector_all("#invBdy > tr")
            house_count = len(rows)
            self.log(f"ENTRIES_DETAILS: Successfully loaded entry page, found {house_count} houses")
            return house_count
        except Exception as exc:
            # Skip this entry gracefully - don't fail the whole process
            self.log(f"ENTRIES_DETAILS: ⚠️ SKIPPING entry (page failed to load): {link}")
            self.log(f"ENTRIES_DETAILS: Error details: {exc}")
            return 0  # Return 0 houses for this entry (skipped)
        finally:
            await page.close()

    # ARCHIVED: No longer used - values extracted from PDF instead
    async def _scrape_single_print7501(self, url: str) -> float:
        """
        Scrape a single print7501 page and return duty sum.
        NOTE: NetCHB site is very slow - using extended timeouts.
        If page fails to load, returns 0.0 (skips this entry gracefully).
        """
        assert self.context
        page = await self.context.new_page()
        try:
            self.log(f"ENTRIES_DETAILS: Loading print7501 page (timeout: 360s): {url}")
            await page.goto(url, wait_until="domcontentloaded", timeout=360000)  # 6 minutes for slow print7501 pages
            self.log(f"ENTRIES_DETAILS: Waiting for print7501 table (timeout: 360s)...")
            await page.wait_for_selector(
                "#pForm > div:nth-child(1) > div:nth-child(2) > div > div.content > table",
                timeout=360000  # 6 minutes for slow print7501 pages
            )
            duty1_elem = await page.query_selector(
                "#pForm > div:nth-child(1) > div:nth-child(2) > div > div.content > table > tbody > tr:nth-child(2) > td:nth-child(2)"
            )
            duty2_elem = await page.query_selector(
                "#pForm > div:nth-child(1) > div:nth-child(2) > div > div.content > table > tbody > tr:nth-child(4) > td:nth-child(2)"
            )
            duty1_text = (await duty1_elem.inner_text()).strip() if duty1_elem else "0"
            duty2_text = (await duty2_elem.inner_text()).strip() if duty2_elem else "0"
            duty_sum = self._parse_currency(duty1_text) + self._parse_currency(duty2_text)
            self.log(f"ENTRIES_DETAILS: Successfully loaded print7501 page, duty sum: {duty_sum:.2f}")
            return duty_sum
        except Exception as exc:
            # Skip this entry gracefully - don't fail the whole process
            self.log(f"ENTRIES_DETAILS: ⚠️ SKIPPING print7501 (page failed to load): {url}")
            self.log(f"ENTRIES_DETAILS: Error details: {exc}")
            return 0.0  # Return 0.0 duty for this entry (skipped)
        finally:
            await page.close()

    async def _process_custom_report(
        self,
        mawb_digits: str,
        oldest_entry: Optional[datetime],
        template_identifier: str,
        format_record: Optional[Dict[str, Any]] = None,
    ) -> Tuple[Optional[Path], Dict[str, str]]:
        """
        Process Custom Report section using HTTP requests only.
        ARCHIVED: Browser method fallback removed to reduce server load.
        
        Args:
            mawb_digits: Normalized MAWB (11 digits)
            oldest_entry: Oldest entry date (required)
            template_identifier: Template identifier (unused, kept for compatibility)
            format_record: Full format record with template_payload (required for HTTP method)
            
        Returns:
            Tuple of (artifact_path, report_summary)
            
        Raises:
            RuntimeError: If template_payload is missing or invalid
        """
        # HTTP method only - no fallback to reduce server load
        if not format_record:
            raise RuntimeError("format_record is required for Custom Report HTTP method")
        
        template_payload = format_record.get("template_payload")
        if not template_payload or not isinstance(template_payload, dict):
            raise RuntimeError("template_payload is required in format_record for Custom Report HTTP method")
        
        header_fields = template_payload.get("headerFields", [])
        manifest_fields = template_payload.get("manifestFields", [])
        if not header_fields or not manifest_fields:
            raise RuntimeError("template_payload must contain headerFields and manifestFields for Custom Report HTTP method")
        
        self.log(f"Using HTTP method with template_payload ({len(header_fields)} headerFields, {len(manifest_fields)} manifestFields)")
        return await self._process_custom_report_http(
            mawb_digits, oldest_entry, template_payload, format_record.get("template_identifier", "")
        )
    
    async def _process_custom_report_http(
        self,
        mawb_digits: str,
        oldest_entry: datetime,
        template_payload: Dict[str, Any],
        template_identifier: str = "",
    ) -> Tuple[Optional[Path], Dict[str, str]]:
        """
        Process Custom Report section using HTTP requests (faster than browser automation).
        
        Args:
            mawb_digits: Normalized MAWB (11 digits)
            oldest_entry: Oldest entry date (required)
            template_payload: Full template configuration from database (must include templateId, headerFields, manifestFields, defaultValues)
            
        Returns:
            Tuple of (artifact_path, report_summary)
        """
        if not self.context:
            raise RuntimeError("Context not initialized")
        
        if not oldest_entry:
            raise RuntimeError("Oldest entry date is required for custom report")
        
        report_summary: Dict[str, str] = {
            "Report Duty": "N/A",
            "Report Total House": "N/A",
            "Total Informal Duty": "N/A",
            "Complete Total Duty": "N/A",
            "Cargo Release Date": "N/A",
        }
        
        download_dir = Path(self.temp_dir.name)
        
        # Get cookies from current browser context
        storage_state = await self.context.storage_state()
        session_cookies = self._load_cookies_from_storage_state(storage_state)
        
        if not session_cookies:
            raise RuntimeError("No cookies found in session - HTTP method requires valid session cookies")
        
        self.log(f"CUSTOM HTTP: Using {len(session_cookies)} cookies from session")
        
        headers = {
            "User-Agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Mobile Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Content-Type": "application/x-www-form-urlencoded",
            "Origin": "https://www.netchb.com",
            "Referer": "https://www.netchb.com/app/entry/customizableReport.jsp",
            "Cache-Control": "max-age=0",
            "Connection": "keep-alive",
            "DNT": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        }
        
        # Build form payload (full template configuration required)
        form_data = self._build_custom_report_payload(template_payload, mawb_digits, oldest_entry)
        self.log(f"CUSTOM HTTP STEP 1: Building payload with templateId=0")
        self.log(f"CUSTOM HTTP STEP 1: Dates: begin={form_data.get('begin')}, end={form_data.get('end')}")
        self.log(f"CUSTOM HTTP STEP 1: MAWB: {form_data.get('masterBill')} (no dash)")
        self.log(f"CUSTOM HTTP STEP 1: Header fields: {len(template_payload.get('headerFields', []))}")
        self.log(f"CUSTOM HTTP STEP 1: Manifest fields: {len(template_payload.get('manifestFields', []))}")
        self.log(f"CUSTOM HTTP STEP 1: Default values: {len(template_payload.get('defaultValues', {}))} fields")
        
        # Custom report generation can take several minutes, so use extended timeout (5 minutes)
        async with httpx.AsyncClient(timeout=300.0, follow_redirects=True) as client:
            # Set cookies
            for name, value in session_cookies.items():
                client.cookies.set(name, value, domain=".netchb.com")
            
            # STEP 1: POST to download endpoint
            self.log("CUSTOM HTTP STEP 2: POST to Custom Report download endpoint...")
            try:
                response = await client.post(
                    CUSTOM_REPORT_DOWNLOAD_URL,
                    data=form_data,
                    headers=headers,
                )
                response.raise_for_status()
                self.log(f"CUSTOM HTTP STEP 2: Response status {response.status_code}")
                self.log(f"CUSTOM HTTP STEP 2: Content-Type: {response.headers.get('content-type')}")
                self.log(f"CUSTOM HTTP STEP 2: Response length: {len(response.content)} bytes")
            except Exception as exc:
                self.log(f"CUSTOM HTTP STEP 2 ERROR: Request failed: {exc}")
                raise RuntimeError(f"CUSTOM HTTP STEP 2 failed: {exc}") from exc
            
            # Check if response is Excel file
            content_type = response.headers.get("content-type", "")
            if "excel" not in content_type.lower() and "spreadsheet" not in content_type.lower():
                self.log(f"CUSTOM HTTP STEP 2 ERROR: Unexpected content type: {content_type}")
                raise RuntimeError(f"Unexpected content type: {content_type}")
            
            # STEP 2: Save Excel file
            filename = f"{mawb_digits[:3]}-{mawb_digits[3:]} customizable report.xlsx"
            file_path = download_dir / filename
            
            self.log(f"CUSTOM HTTP STEP 3: Saving Excel file to: {file_path}")
            try:
                with open(file_path, "wb") as f:
                    f.write(response.content)
                self.log(f"CUSTOM HTTP STEP 3: ✓ File saved successfully ({len(response.content)} bytes)")
            except Exception as exc:
                self.log(f"CUSTOM HTTP STEP 3 ERROR: Failed to save file: {exc}")
                raise
            
            # STEP 3: Parse Excel file
            # Use template_identifier to determine parser (since we removed templateId)
            self.log("CUSTOM HTTP STEP 4: Parsing Excel file...")
            try:
                report_summary.update(self._parse_custom_report_excel(file_path, template_identifier=template_identifier))
                self.log("CUSTOM HTTP STEP 4: ✓ Excel parsed successfully")
                self.log(f"CUSTOM HTTP STEP 4: Report Duty: {report_summary.get('Report Duty')}")
                self.log(f"CUSTOM HTTP STEP 4: Report Total House: {report_summary.get('Report Total House')}")
            except Exception as exc:
                self.log(f"CUSTOM HTTP STEP 4 ERROR: Failed to parse Excel: {exc}")
                raise
            
            # STEP 4: Upload to AWS S3
            # Note: This is an optional early upload. Main upload happens in service.py with full info (airport_code, customer, template_name)
            self.log("CUSTOM HTTP STEP 5: Uploading Excel to AWS S3...")
            try:
                from .storage import NetChbDutyStorageManager
                storage_manager = NetChbDutyStorageManager()
                # Extract template name from template_identifier for V2 suffix detection
                template_name = None
                if template_identifier and "shoaib" in template_identifier.lower():
                    template_name = "Shoaib Match"  # Use known template name for V2 detection
                storage_path, signed_url = storage_manager.upload_excel(
                    file_path,
                    mawb_digits,
                    airport_code=None,  # Not available in this context
                    customer=None,  # Not available in this context
                    template_name=template_name,
                )
                report_summary["excel_storage_path"] = storage_path
                report_summary["excel_download_url"] = signed_url
                self.log(f"CUSTOM HTTP STEP 5: ✓ Excel uploaded to storage: {storage_path}")
            except Exception as exc:
                self.log(f"CUSTOM HTTP STEP 5 WARNING: Failed to upload to storage: {exc} (continuing without storage)")
                # Don't fail the entire process if storage upload fails
            
            self.log("CUSTOM HTTP: Custom Report section complete (HTTP method)")
            return file_path, report_summary
    
    async def _process_custom_report_browser(
        self,
        mawb_digits: str,
        oldest_entry: Optional[datetime],
        template_identifier: str,
    ) -> Tuple[Optional[Path], Dict[str, str]]:
        """
        ARCHIVED: Process Custom Report section using browser automation (no longer used - fallback removed).
        """
        assert self.page
        report_summary: Dict[str, str] = {
            "Report Duty": "N/A",
            "Report Total House": "N/A",
            "Total Informal Duty": "N/A",
            "Complete Total Duty": "N/A",
            "Cargo Release Date": "N/A",
        }

        download_dir = Path(self.temp_dir.name)
        self.log("CUSTOM STEP 1: Opening Customizable Report page in new tab...")
        report_page = await self.context.new_page()
        await report_page.goto(CUSTOM_REPORT_URL, wait_until="domcontentloaded", timeout=30000)
        self.log(f"CUSTOM STEP 1: Report page loaded. URL: {report_page.url}")

        self.log("CUSTOM STEP 2: Waiting for template selector (#sTemp)...")
        await report_page.wait_for_selector("#sTemp", timeout=30000)
        self.log("CUSTOM STEP 2: Template selector found")

        self.log(f"CUSTOM STEP 3: Searching for template: {template_identifier}")
        template_found = False
        options = await report_page.query_selector_all("#sTemp option")
        for option in options:
            text = await option.inner_text()
            if template_identifier.lower() in text.lower():
                await report_page.select_option("#sTemp", value=await option.get_attribute("value"))
                template_found = True
                self.log(f"CUSTOM STEP 3: Template selected: {text}")
                break

        if not template_found:
            raise RuntimeError(f"Template '{template_identifier}' not found in dropdown")

        if not oldest_entry:
            raise RuntimeError("Oldest entry date is required for custom report")

        self.log(f"CUSTOM STEP 4: Filling entry date: {oldest_entry.strftime('%m%d%y')}")
        await report_page.fill("#rd", oldest_entry.strftime("%m%d%y"))
        self.log("CUSTOM STEP 4: Entry date filled")

        # Optimize end date: If entry date is 25+ days old, use 25th day as end date
        # This prevents very slow queries when entry date is very old
        today = datetime.now()
        days_old = (today - oldest_entry).days
        
        if days_old >= 25:
            # Use 25th day from entry date as end date (optimize query performance)
            end_date = oldest_entry + timedelta(days=25)
            self.log(f"CUSTOM STEP 5: Entry date is {days_old} days old, using end date = entry date + 25 days ({end_date.strftime('%m/%d/%y')})")
        else:
            # Use today's date as end date (normal case)
            end_date = today
            self.log(f"CUSTOM STEP 5: Using today's date as end date: {end_date.strftime('%m/%d/%y')}")
        
        await report_page.fill("#le", end_date.strftime("%m%d%y"))
        self.log("CUSTOM STEP 5: End date filled")

        # Format MAWB as XXX-XXXXXXXX (with dash)
        mawb_clean = "".join(c for c in mawb_digits if c.isdigit())
        if len(mawb_clean) == 11:
            formatted_mawb = f"{mawb_clean[:3]}-{mawb_clean[3:]}"
        else:
            formatted_mawb = mawb_clean
        
        self.log(f"CUSTOM STEP 6: Filling MAWB: {formatted_mawb}")
        await report_page.fill("#mbl", formatted_mawb)
        self.log("CUSTOM STEP 6: MAWB filled")

        self.log("CUSTOM STEP 7: Selecting 'All Users'...")
        await report_page.select_option("#usr", label="All Users")
        self.log("CUSTOM STEP 7: User selected")

        self.log("CUSTOM STEP 8: Selecting 'All Locations'...")
        await report_page.select_option("#loc", label="All Locations")
        self.log("CUSTOM STEP 8: Location selected")

        before_download = set(download_dir.glob("*"))
        self.log("CUSTOM STEP 9: Clicking download button (#drB)...")
        
        async with report_page.expect_download(timeout=120000) as download_info:
            await report_page.click("#drB")
        
        download = await download_info.value
        self.log(f"CUSTOM STEP 9: Download started: {download.suggested_filename}")
        
        downloaded_file = download_dir / download.suggested_filename
        await download.save_as(downloaded_file)
        self.log(f"CUSTOM STEP 9: Download saved to: {downloaded_file}")

        renamed = download_dir / f"{mawb_digits[:3]}-{mawb_digits[3:]} customizable report.xlsx"
        if downloaded_file != renamed:
            shutil.move(str(downloaded_file), renamed)
            self.log(f"CUSTOM STEP 10: File renamed to: {renamed}")

        # Use template_identifier for parsing (browser method)
        report_summary.update(self._parse_custom_report_excel(renamed, template_identifier=template_identifier))
        self.log("CUSTOM STEP 11: Excel file parsed")

        await report_page.close()
        self.log("CUSTOM STEP 12: Report tab closed")

        return renamed, report_summary

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _parse_custom_report_excel(self, path: Path, template_identifier: Optional[str] = None) -> Dict[str, str]:
        """
        Parse custom report Excel file based on template type.
        
        Args:
            path: Path to Excel file
            template_identifier: Template identifier to determine parsing logic (e.g., "fte-match", "shoaib-match")
        """
        if template_identifier and "shoaib" in template_identifier.lower():
            return self._parse_shoaib_match_excel(path)
        elif template_identifier and "fte" in template_identifier.lower():
            return self._parse_fte_match_excel(path)
        else:
            # Default to FTE Match parsing for backward compatibility
            self.log(f"PARSING EXCEL: Unknown template_identifier={template_identifier}, using FTE Match parser")
            return self._parse_fte_match_excel(path)
    
    def _parse_fte_match_excel(self, path: Path) -> Dict[str, str]:
        """Parse FTE Match template Excel file (templateId: 3351)."""
        self.log(f"PARSING EXCEL (FTE Match): Reading {path}")
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        total_duty = 0.0
        total_house = 0
        total_informal = 0.0
        complete_duty = 0.0
        entry_dates = set()
        release_dates = set()

        for row in sheet.iter_rows(min_row=2):
            try:
                informal = row[4].value or 0
                complete = row[6].value or 0
                entry_dates_val = row[2].value
                release_date_val = row[8].value
                if row[13].value not in (None, ""):
                    total_house += 1
                total_informal += float(informal)
                complete_duty += float(complete)
                total_duty += float(informal) + float(complete)
                if entry_dates_val not in (None, ""):
                    entry_dates.add(_format_excel_date(entry_dates_val))
                if release_date_val not in (None, ""):
                    release_dates.add(_format_excel_date(release_date_val))
            except Exception:
                continue

        self.log(f"PARSING EXCEL (FTE Match): Total duty={total_duty:.2f}, Houses={total_house}, Informal={total_informal:.2f}, Complete={complete_duty:.2f}")

        return {
            "Report Duty": f"{total_duty:.2f}",
            "Report Total House": str(total_house),
            "Total Informal Duty": f"{total_informal:.2f}",
            "Complete Total Duty": f"{complete_duty:.2f}",
            "Entry Date": ", ".join(sorted(entry_dates)) or "N/A",
            "Cargo Release Date": ", ".join(sorted(release_dates)) or "N/A",
        }
    
    def _parse_shoaib_match_excel(self, path: Path) -> Dict[str, str]:
        """
        Parse Shoaib Match template Excel file (templateId: 2836).
        
        Column mapping:
        - Column A (index 0): Entry identifier (used for deduplication)
        - Column D (index 3): Entry date
        - Column F (index 5): Informal duty
        - Column H (index 7): Complete duty
        - Column J (index 9): Cargo release date
        - Column N (index 13): House indicator (ignore empty cells and header)
        
        Logic:
        - Deduplicate duties by Column A: Sum Column F (Informal) and Column H (Complete) only for unique Column A values
        - Count houses from ALL rows: Count Column N for every row (ignore empty cells, but count all non-empty)
        - Extract unique dates from Column D and Column J
        """
        self.log(f"PARSING EXCEL (Shoaib Match): Reading {path}")
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        total_duty = 0.0
        total_house = 0
        total_informal = 0.0
        complete_duty = 0.0
        entry_dates = set()
        release_dates = set()
        
        # Track unique Column A values for duty deduplication
        unique_entries = {}  # {column_a_value: (informal, complete)}
        
        # Process all rows in a single pass
        for row in sheet.iter_rows(min_row=2):
            try:
                column_a = row[0].value
                column_d = row[3].value  # Entry date
                column_f = row[5].value or 0  # Informal duty
                column_h = row[7].value or 0  # Complete duty
                column_j = row[9].value  # Release date
                column_n = row[13].value  # House indicator
                
                # Skip if Column A is empty or None (header row or invalid)
                if column_a is None or column_a == "":
                    continue
                
                # Count houses from ALL rows (not deduplicated by Column A)
                # Ignore empty cells in Column N
                if column_n not in (None, ""):
                    total_house += 1
                
                # Deduplicate duties by Column A (only sum once per unique Column A value)
                if column_a not in unique_entries:
                    unique_entries[column_a] = (float(column_f), float(column_h))
                    total_informal += float(column_f)
                    complete_duty += float(column_h)
                
                # Collect dates (from all rows, not just unique)
                if column_d not in (None, ""):
                    entry_dates.add(_format_excel_date(column_d))
                if column_j not in (None, ""):
                    release_dates.add(_format_excel_date(column_j))
            except Exception:
                continue
        
        total_duty = total_informal + complete_duty
        
        self.log(f"PARSING EXCEL (Shoaib Match): Total duty={total_duty:.2f}, Houses={total_house}, Informal={total_informal:.2f}, Complete={complete_duty:.2f}, Unique entries={len(unique_entries)}")

        return {
            "Report Duty": f"{total_duty:.2f}",
            "Report Total House": str(total_house),
            "Total Informal Duty": f"{total_informal:.2f}",
            "Complete Total Duty": f"{complete_duty:.2f}",
            "Entry Date": ", ".join(sorted(entry_dates)) or "N/A",
            "Cargo Release Date": ", ".join(sorted(release_dates)) or "N/A",
        }

    def _build_custom_report_payload(
        self,
        template_payload: Dict[str, Any],
        mawb_digits: str,
        oldest_entry: datetime
    ) -> Dict[str, Any]:
        """
        Build the form payload for custom report download request.
        
        Uses templateId = "0" while keeping all other fields from template_payload.
        NetCHB accepts templateId = "0" and uses headerFields/manifestFields to identify the template.
        
        Args:
            template_payload: Full template configuration from database (must include headerFields, manifestFields, defaultValues)
            mawb_digits: Full MAWB (11 digits)
            oldest_entry: Oldest entry date (datetime object)
            
        Returns:
            Dictionary of form data for POST request with all template fields and templateId = "0".
        """
        payload = {}
        
        # Template ID = "0" (tested and confirmed working - NetCHB accepts this value)
        payload["templateId"] = "0"
        
        # Dates (MMDDYY format)
        # Optimize end date: If entry date is 25+ days old, use 25th day as end date
        # This prevents very slow queries when entry date is very old
        today = datetime.now()
        days_old = (today - oldest_entry).days
        
        if days_old >= 25:
            # Use 25th day from entry date as end date (optimize query performance)
            end_date = oldest_entry + timedelta(days=25)
            self.log(f"CUSTOM REPORT DATE OPTIMIZATION: Entry date is {days_old} days old, using end date = entry date + 25 days ({end_date.strftime('%m/%d/%y')})")
        else:
            # Use today's date as end date (normal case)
            end_date = today
        
        payload["begin"] = oldest_entry.strftime("%m%d%y")
        payload["end"] = end_date.strftime("%m%d%y")
        
        # MAWB - Send WITHOUT dash (just digits) - matching network capture format
        # Remove any existing dashes/spaces and send as plain digits
        mawb_clean = "".join(c for c in mawb_digits if c.isdigit())
        payload["masterBill"] = mawb_clean  # No dash formatting
        
        # Default values from template (all filter fields, status fields, etc.)
        default_values = template_payload.get("defaultValues", {})
        payload.update(default_values)
        
        # Header fields (array - httpx handles lists by creating multiple form fields with same key)
        header_fields = template_payload.get("headerFields", [])
        if header_fields:
            payload["headerFields"] = header_fields
        
        # Manifest fields (array - httpx handles lists by creating multiple form fields with same key)
        manifest_fields = template_payload.get("manifestFields", [])
        if manifest_fields:
            payload["manifestFields"] = manifest_fields
        
        # Invoice fields (array - for templates that include invoice data)
        invoice_fields = template_payload.get("invoiceFields", [])
        if invoice_fields:
            payload["invoiceFields"] = invoice_fields
        
        # Line fields (array - for templates that include line item data)
        line_fields = template_payload.get("lineFields", [])
        if line_fields:
            payload["lineFields"] = line_fields
        
        # Tariff fields (array - for templates that include tariff data)
        tariff_fields = template_payload.get("tariffFields", [])
        if tariff_fields:
            payload["tariffFields"] = tariff_fields
        
        return payload
    
    @staticmethod
    def _parse_currency(value: str) -> float:
        try:
            return float(value.replace(",", "").replace("$", ""))
        except Exception:
            return 0.0

    async def _download_7501_batch_pdf_http(
        self,
        mawb_digits: str,
        entry_rows: Optional[List[Dict]] = None,
        session_cookies: Optional[Dict[str, str]] = None,
    ) -> Optional[Path]:
        """
        Download 7501 batch PDF for verified masters using HTTP requests.
        
        Args:
            mawb_digits: Normalized MAWB (11 digits)
            entry_rows: Optional pre-fetched entry rows (to avoid redundant requests)
            session_cookies: Optional session cookies (if not provided, uses context cookies)
            
        Returns:
            Path to downloaded and compressed PDF file, or None if failed
        """
        if not self.context:
            raise RuntimeError("Context not initialized")
        
        # Get cookies if not provided
        if not session_cookies:
            storage_state = await self.context.storage_state()
            session_cookies = self._load_cookies_from_storage_state(storage_state)
        
        if not session_cookies:
            self.log("⚠ No cookies found for PDF download")
            return None
        
        self.log(f"PDF DOWNLOAD: Starting 7501 batch PDF download for MAWB {mawb_digits}")
        
        headers = {
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 18_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.5 Mobile/15E148 Safari/604.1",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Content-Type": "application/x-www-form-urlencoded",
            "Origin": "https://www.netchb.com",
            "Referer": "https://www.netchb.com/app/entry/index.jsp",
            "Cache-Control": "max-age=0",
            "Connection": "keep-alive",
            "DNT": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        }
        
        # PDF generation can take several minutes for large batches, so use extended timeout (10 minutes)
        # Note: This timeout is necessary as NetCHB can take time to generate PDFs
        async with httpx.AsyncClient(timeout=600.0, follow_redirects=True) as client:
            # Set cookies
            for name, value in session_cookies.items():
                client.cookies.set(name, value, domain=".netchb.com")
            
            # STEP 1: Get entry numbers (reuse if provided, otherwise fetch)
            entry_numbers = []
            
            def extract_entry_no_from_row(row):
                """Extract entry number from entry row (from query_string or link)."""
                # Try query_string first (format: "filerCode=...&entryNo=12345")
                query_string = row.get("query_string", "")
                if query_string:
                    match = re.search(r"entryNo=(\d+)", query_string)
                    if match:
                        return match.group(1)
                
                # Fallback: try link (format: "?filerCode=...&entryNo=12345")
                link = row.get("link", "")
                if link:
                    match = re.search(r"entryNo=(\d+)", link)
                    if match:
                        return match.group(1)
                
                return None
            
            if entry_rows:
                self.log(f"PDF DOWNLOAD STEP 1: Using pre-fetched entry_rows ({len(entry_rows)} rows)")
                entry_numbers = []
                for row in entry_rows:
                    entry_no = extract_entry_no_from_row(row)
                    if entry_no:
                        entry_numbers.append(entry_no)
                self.log(f"PDF DOWNLOAD STEP 1: Extracted {len(entry_numbers)} entry numbers from pre-fetched rows")
            else:
                self.log("PDF DOWNLOAD STEP 1: Fetching entries data...")
                form_data = {
                    "entryNoSearch": "",
                    "brokerRefNo": "",
                    "importerRecord": "0",
                    "importerRecordName": "",
                    "importerSearchByProfile": "true",
                    "ultimateConsignee": "0",
                    "ultimateConsigneeName": "",
                    "ultimateConsigneeSearchByProfile": "true",
                    "freightForwarder": "0",
                    "freightForwarderName": "",
                    "freightForwarderSearchByProfile": "true",
                    "begin": "",
                    "end": "",
                    "entryStatus": "",
                    "cargoReleaseStatus": "",
                    "manifestStatus": "",
                    "pgaAgency": "",
                    "ogaStatus": "",
                    "statusColor": "",
                    "entryType": "",
                    "portEntry": "",
                    "modeTransport": "",
                    "masterBill": mawb_digits,
                    "searchTimePeriod": "Y1",
                    "user": "",
                    "location": "0",
                    "noPerPage": "1000",
                    "entryNo": "0",
                    "orderBy": "vep1",
                    "page": "0",
                    "unchecked7501": "",
                    "unchecked3461": "",
                    "method": "view",
                }
                
                try:
                    response = await client.post(
                        ENTRIES_SEARCH_POST_URL,
                        data=form_data,
                        headers=headers,
                    )
                    response.raise_for_status()
                    search_data = self._parse_entries_search_results(response.text)
                    if search_data and search_data.get("entry_rows"):
                        for row in search_data["entry_rows"]:
                            entry_no = extract_entry_no_from_row(row)
                            if entry_no:
                                entry_numbers.append(entry_no)
                except Exception as exc:
                    self.log(f"PDF DOWNLOAD STEP 1 ERROR: Failed to fetch entries: {exc}")
                    return None
            
            if not entry_numbers:
                self.log("PDF DOWNLOAD STEP 1: No entries found for this MAWB")
                return None
            
            self.log(f"PDF DOWNLOAD STEP 1: Found {len(entry_numbers)} entries")
            
            # STEP 2: DIRECT PDF GENERATION (skipping form page request)
            # Tested and confirmed: We can generate PDF directly without requesting form page first
            # This saves ~10 seconds per PDF download by eliminating one HTTP request
            # 
            # BACKUP CODE (commented for reference - in case direct method fails in future):
            # ============================================================================
            # # STEP 2: Request PDF form page
            # self.log("PDF DOWNLOAD STEP 2: Requesting PDF form page...")
            # pdf_form_payload = {
            #     "entryNoSearch": "",
            #     "brokerRefNo": "",
            #     "importerRecord": "0",
            #     "importerRecordName": "",
            #     "importerSearchByProfile": "true",
            #     "ultimateConsignee": "0",
            #     "ultimateConsigneeName": "",
            #     "ultimateConsigneeSearchByProfile": "true",
            #     "freightForwarder": "0",
            #     "freightForwarderName": "",
            #     "freightForwarderSearchByProfile": "true",
            #     "begin": "",
            #     "end": "",
            #     "entryStatus": "",
            #     "cargoReleaseStatus": "",
            #     "manifestStatus": "",
            #     "pgaAgency": "",
            #     "ogaStatus": "",
            #     "statusColor": "",
            #     "entryType": "",
            #     "portEntry": "",
            #     "modeTransport": "",
            #     "masterBill": mawb_digits,
            #     "searchTimePeriod": "Y1",
            #     "user": "",
            #     "location": "0",
            #     "noPerPage": "1000",
            #     "entryNo": "0",
            #     "orderBy": "vep1",
            #     "page": "0",
            #     "unchecked7501": "",
            #     "unchecked3461": "",
            #     "method": "print7501Batch",
            # }
            # 
            # # Add print7501[ENTRY_NO]=true for each entry
            # for entry_no in entry_numbers:
            #     pdf_form_payload[f"print7501[{entry_no}]"] = "true"
            # 
            # try:
            #     form_response = await client.post(
            #         ENTRIES_SEARCH_POST_URL,
            #         data=pdf_form_payload,
            #         headers=headers,
            #     )
            #     form_response.raise_for_status()
            #     
            #     # Parse PDF form page
            #     soup = BeautifulSoup(form_response.text, "html.parser")
            #     form = soup.find("form", action="/app/entry/7501_Batch.pdf")
            #     if not form:
            #         self.log("PDF DOWNLOAD STEP 2 ERROR: PDF form not found")
            #         return None
            #     
            #     entry_nos_input = form.find("input", {"name": "entryNos"})
            #     entry_nos_value = entry_nos_input.get("value", "") if entry_nos_input else ""
            #     
            #     type_input = form.find("input", {"name": "type"})
            #     type_value = type_input.get("value", "6") if type_input else "6"
            #     
            #     signature_input = form.find("input", {"name": "signature"})
            #     signature_value = signature_input.get("value", "") if signature_input else ""
            #     
            #     date_input = form.find("input", {"name": "signedDate"})
            #     signed_date = datetime.now().strftime("%m%d%y")
            #     if date_input:
            #         signed_date = date_input.get("value", signed_date)
            #     
            #     self.log(f"PDF DOWNLOAD STEP 2: ✓ PDF form page loaded (entryNos: {entry_nos_value[:50]}...)")
            #     
            # except Exception as exc:
            #     self.log(f"PDF DOWNLOAD STEP 2 ERROR: Failed to get PDF form: {exc}")
            #     return None
            # ============================================================================
            
            # Direct PDF generation - construct payload directly without form page
            self.log("PDF DOWNLOAD STEP 2: Constructing PDF payload directly (skipping form page request)...")
            
            # Construct entryNos value: comma-separated entry numbers with trailing comma
            # Format: "12345,67890,11111,"
            entry_nos_value = ",".join(entry_numbers) + ","
            
            # Set default values directly (no need to fetch from form)
            type_value = "6"  # Always "6" based on testing
            signature_value = ""  # Empty/default
            signed_date = datetime.now().strftime("%m%d%y")  # Current date in MMDDYY format
            
            self.log(f"PDF DOWNLOAD STEP 2: ✓ Payload constructed directly (entryNos: {entry_nos_value[:50]}..., type: {type_value}, date: {signed_date})")
            
            # STEP 3: Generate PDF
            self.log(f"PDF DOWNLOAD STEP 3: Generating PDF for {len(entry_numbers)} entries...")
            self.log("PDF DOWNLOAD STEP 3: ⏳ This may take several minutes for large batches. Please wait...")
            PDF_BATCH_URL = "https://www.netchb.com/app/entry/7501_Batch.pdf"
            
            pdf_generation_payload = {
                "signature": signature_value,
                "digitalSignature": "",
                "signedDate": signed_date,
                "broker": "false",
                "cashier": "false",
                "record": "false",
                "original": "false",
                "multiple": "false",
                "type7501": "2",  # New 7501 Format
                "separateConsignees": "false",
                "printPartNumbers": "false",
                "printMfrName": "false",
                "entryNoBlank": "false",
                "entryNos": entry_nos_value,
                "type": type_value,
            }
            
            try:
                # PDF generation request with extended timeout (30 minutes)
                # NetCHB can take a very long time to generate batch PDFs
                import time
                start_time = time.time()
                
                pdf_response = await client.post(
                    PDF_BATCH_URL,
                    data=pdf_generation_payload,
                    headers={**headers, "Content-Type": "application/x-www-form-urlencoded"},
                )
                
                elapsed_time = time.time() - start_time
                self.log(f"PDF DOWNLOAD STEP 3: ⏱️ PDF generation request completed in {elapsed_time:.1f} seconds ({elapsed_time/60:.1f} minutes)")
                
                pdf_response.raise_for_status()
                
                # Check if response is PDF
                content_type = pdf_response.headers.get("content-type", "")
                if "pdf" not in content_type.lower():
                    self.log(f"PDF DOWNLOAD STEP 3 ERROR: Unexpected content type: {content_type}")
                    self.log(f"PDF DOWNLOAD STEP 3 ERROR: Response length: {len(pdf_response.content)} bytes")
                    # Sometimes NetCHB returns HTML error pages instead of PDF
                    if len(pdf_response.content) < 10000:  # Small response likely an error page
                        try:
                            error_text = pdf_response.text[:500]
                            self.log(f"PDF DOWNLOAD STEP 3 ERROR: Response preview: {error_text}")
                        except:
                            pass
                    return None
                
                # Save original PDF
                download_dir = Path(self.temp_dir.name)
                original_pdf_path = download_dir / f"{mawb_digits}_7501_batch_original.pdf"
                with open(original_pdf_path, "wb") as f:
                    f.write(pdf_response.content)
                
                original_size = len(pdf_response.content)
                self.log(f"PDF DOWNLOAD STEP 3: ✓ PDF downloaded successfully ({original_size:,} bytes, {original_size/1024/1024:.2f} MB)")
                
                # STEP 4: Compress PDF
                self.log("PDF DOWNLOAD STEP 4: Compressing PDF...")
                compressed_pdf_path = download_dir / f"{mawb_digits}_7501_batch.pdf"
                
                try:
                    compressed_pdf_path = self._compress_pdf_ghostscript(original_pdf_path, compressed_pdf_path)
                    compressed_size = compressed_pdf_path.stat().st_size
                    reduction_pct = ((original_size - compressed_size) / original_size) * 100
                    self.log(f"PDF DOWNLOAD STEP 4: ✓ PDF compressed ({original_size:,} bytes → {compressed_size:,} bytes, {reduction_pct:.1f}% reduction)")
                    
                    # Delete original PDF
                    original_pdf_path.unlink()
                    
                    return compressed_pdf_path
                    
                except Exception as exc:
                    self.log(f"PDF DOWNLOAD STEP 4 WARNING: Compression failed: {exc} - using original PDF")
                    # Fallback to original if compression fails
                    compressed_pdf_path.unlink(missing_ok=True)
                    original_pdf_path.rename(compressed_pdf_path)
                    return compressed_pdf_path
                
            except Exception as exc:
                self.log(f"PDF DOWNLOAD STEP 3 ERROR: Failed to generate PDF: {exc}")
                return None

    def _compress_pdf_ghostscript(self, input_path: Path, output_path: Path) -> Path:
        """
        Compress PDF aggressively using Ghostscript (target: 97-98% reduction).
        
        Args:
            input_path: Path to original PDF
            output_path: Path to save compressed PDF
            
        Returns:
            Path to compressed PDF
        """
        import subprocess
        
        self.log(f"PDF COMPRESSION: Starting Ghostscript compression...")
        
        # Ghostscript command for aggressive compression
        gs_command = [
            "gs",
            "-sDEVICE=pdfwrite",
            "-dCompatibilityLevel=1.4",
            "-dPDFSETTINGS=/screen",  # Aggressive compression
            "-dNOPAUSE",
            "-dQUIET",
            "-dBATCH",
            "-dColorImageResolution=150",  # Reduce image resolution
            "-dGrayImageResolution=150",
            "-dMonoImageResolution=150",
            "-dColorImageDownsampleType=/Bicubic",
            "-dGrayImageDownsampleType=/Bicubic",
            "-dColorConversionStrategy=/sRGB",
            "-dProcessColorModel=/DeviceRGB",
            f"-sOutputFile={output_path}",
            str(input_path),
        ]
        
        try:
            result = subprocess.run(
                gs_command,
                capture_output=True,
                text=True,
                timeout=120,  # 2 minutes timeout
            )
            
            if result.returncode != 0:
                raise RuntimeError(f"Ghostscript compression failed: {result.stderr}")
            
            # Log compression ratio
            original_size = input_path.stat().st_size
            compressed_size = output_path.stat().st_size
            reduction_pct = ((original_size - compressed_size) / original_size) * 100
            
            self.log(f"PDF COMPRESSION: ✓ Compressed {original_size:,} bytes → {compressed_size:,} bytes ({reduction_pct:.1f}% reduction)")
            
            return output_path
            
        except FileNotFoundError:
            raise RuntimeError("Ghostscript not found. Please install Ghostscript: brew install ghostscript (macOS) or apt-get install ghostscript (Linux)")
        except subprocess.TimeoutExpired:
            raise RuntimeError("Ghostscript compression timed out after 2 minutes")
        except Exception as exc:
            raise RuntimeError(f"Ghostscript compression failed: {str(exc)}")

